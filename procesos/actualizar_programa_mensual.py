from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

from procesos.clasificar_cierre_ot import cargar_diccionarios
from utils.excel_utils import save_dataframe_to_excel
from utils.texto_utils import (
    first_non_empty,
    normalize_column_name,
    normalize_key,
    normalize_text,
    safe_str,
)

SEMANAL_ESTADO_COL_HINTS = {
    "R",
    "NR",
    "RR",
    "S",
    "REALIZADO",
    "NO_REALIZADO",
}

MENSUAL_CODIGO_HEADERS_PRIORITY = [
    "CODIGO",
    "CODIGO_MST",
    "CODIGO_PUNTO",
]
MENSUAL_MATRIZ_TERRENO_HEADER = "MATRIZ_TERRENO"
ESTADOS_PROTEGIDOS = {"R", "NR", "RR", "S"}

MESES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

MATRIZ_ALIAS_CANONICAL: Dict[str, set[str]] = {
    "AP": {"AP", "AGUA POTABLE"},
    "AR": {"AR", "AGUA RESIDUAL"},
    "ASUB": {
        "ASUB",
        "AGUA SUBTERRANEA",
        "AGUAS SUBTERRANEAS",
        "AGUA SUBTERRANEA S",
    },
    "NF": {
        "NF",
        "NIVEL FREATICO",
        "NIVELES FREATICOS",
    },
    "ASUP": {
        "ASUP",
        "AGUA SUPERFICIAL",
        "AGUAS SUPERFICIALES",
    },
    "CAUDAL": {
        "CAUDAL",
        "CAUDALES",
    },
    "FOTOMETRO": {
        "FOTOMETRO",
        "ESTACION 14",
        "ESTACION14",
    },
    "HK": {"HK", "HOUSEKEEPING"},
}

MATRIZ_ALIAS_CANONICAL_NORMALIZED: Dict[str, set[str]] = {
    canonical: {normalize_text(v) for v in aliases}
    for canonical, aliases in MATRIZ_ALIAS_CANONICAL.items()
}


def _canonical_matriz(value: object) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    for canonical, normalized_aliases in MATRIZ_ALIAS_CANONICAL_NORMALIZED.items():
        if normalized in normalized_aliases:
            return canonical
    return normalized


def _safe_sheet_limits(ws) -> tuple[int, int]:
    max_row = ws.max_row if isinstance(ws.max_row, int) and ws.max_row >= 1 else 1
    max_col = ws.max_column if isinstance(ws.max_column, int) and ws.max_column >= 1 else 1
    return max_row, max_col


def _extract_day(value: object) -> int | None:
    if isinstance(value, int):
        return value if 1 <= value <= 31 else None
    if isinstance(value, float) and value.is_integer():
        day = int(value)
        return day if 1 <= day <= 31 else None
    if hasattr(value, "day"):
        try:
            day = int(value.day)
            return day if 1 <= day <= 31 else None
        except Exception:
            return None
    value_str = safe_str(value)
    if value_str.isdigit():
        day = int(value_str)
        return day if 1 <= day <= 31 else None
    return None


def _extract_date(value: object) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def _is_estado_header(normalized_header: str) -> bool:
    if not normalized_header:
        return False
    if normalized_header in SEMANAL_ESTADO_COL_HINTS:
        return True
    if "ESTADO" in normalized_header or "STATUS" in normalized_header:
        return True
    return "REALIZADO" in normalized_header


def _select_sheet_by_name(wb, preferred_name: str):
    preferred_norm = normalize_text(preferred_name)
    for ws in wb.worksheets:
        if normalize_text(ws.title) == preferred_norm:
            return ws
    return None


def _detectar_layout_semanal(ws) -> tuple[int, int, int, List[int]]:
    max_row, max_col = _safe_sheet_limits(ws)
    best_date_row = None
    best_day_cols: List[int] = []

    for row in range(1, min(25, max_row) + 1):
        day_cols: List[int] = []
        for col in range(1, max_col + 1):
            if _extract_day(ws.cell(row=row, column=col).value) is not None:
                day_cols.append(col)
        if len(day_cols) > len(best_day_cols):
            best_day_cols = day_cols
            best_date_row = row

    if best_date_row is None or len(best_day_cols) < 2:
        raise ValueError(
            f"No se detecto una fila de fechas/dias valida en hoja semanal '{ws.title}'."
        )

    header_row = max(1, best_date_row - 1)
    data_start_row = best_date_row + 1
    return header_row, best_date_row, data_start_row, best_day_cols


def _detectar_bloques_dia(
    ws, header_row: int, date_row: int, day_cols: List[int]
) -> List[Tuple[int, dt.date | None, int, int, int | None, int]]:
    """
    Retorna bloques diarios en formato:
    (dia, fecha_ref, col_codigo, col_actividad, col_tarea, col_estado)
    """
    _, max_col = _safe_sheet_limits(ws)
    blocks: List[Tuple[int, dt.date | None, int, int, int | None, int]] = []

    for idx, day_col in enumerate(day_cols):
        day_value = ws.cell(row=date_row, column=day_col).value
        day = _extract_day(day_value)
        if day is None:
            continue
        fecha_ref = _extract_date(day_value)

        next_day_col = day_cols[idx + 1] if idx + 1 < len(day_cols) else (max_col + 1)
        search_end = min(next_day_col - 1, day_col + 8, max_col)

        actividad_col = None
        tarea_col = None
        estado_col = None

        for col in range(day_col, search_end + 1):
            header = normalize_column_name(ws.cell(row=header_row, column=col).value)
            if not header:
                continue
            if actividad_col is None and "ACTIVIDAD" in header:
                actividad_col = col
                if "TAREA" in header:
                    tarea_col = col
                continue
            if tarea_col is None and "TAREA" in header:
                tarea_col = col
                continue
            if estado_col is None and _is_estado_header(header):
                estado_col = col

        if actividad_col is None or estado_col is None:
            continue

        blocks.append((day, fecha_ref, day_col, actividad_col, tarea_col, estado_col))

    return blocks


def _leer_semanal_horizontal(
    path_programa_turno: Path | str,
    hoja_preferida: str = "Control SGS",
) -> tuple[pd.DataFrame, str]:
    print("[3/5] Leyendo programa semanal (horizontal)...")
    wb = load_workbook(Path(path_programa_turno), data_only=True, read_only=True)
    records: List[Dict[str, object]] = []
    ws = None
    try:
        ws = _select_sheet_by_name(wb, hoja_preferida)
        if ws is None:
            disponibles = ", ".join(wb.sheetnames)
            raise ValueError(
                f"No se encontro la hoja semanal requerida '{hoja_preferida}'. "
                f"Hojas disponibles: {disponibles}"
            )

        header_row, date_row, data_start_row, day_cols = _detectar_layout_semanal(ws)
        blocks = _detectar_bloques_dia(ws, header_row, date_row, day_cols)
        if not blocks:
            raise ValueError(
                f"No se detectaron bloques diarios validos en hoja semanal '{ws.title}'."
            )

        max_row, _ = _safe_sheet_limits(ws)
        for row in range(data_start_row, max_row + 1):
            for day, fecha_ref, col_codigo, col_actividad, col_tarea, col_estado in blocks:
                codigo = safe_str(ws.cell(row=row, column=col_codigo).value)
                actividad = safe_str(ws.cell(row=row, column=col_actividad).value)
                tarea = safe_str(ws.cell(row=row, column=col_tarea).value) if col_tarea else ""
                estado = safe_str(ws.cell(row=row, column=col_estado).value)

                if not any([actividad, tarea, estado]):
                    continue
                if normalize_text(codigo) in {"TOTAL", "SUBTOTAL"}:
                    continue

                records.append(
                    {
                        "CODIGO": codigo,
                        "ACTIVIDAD": actividad,
                        "TAREA": tarea,
                        "ACTIVIDAD_DICCIONARIO": first_non_empty([actividad, tarea]),
                        "DIA": day,
                        "ESTADO": estado,
                        "FECHA_REFERENCIA": fecha_ref,
                        "ORIGEN_HOJA": ws.title,
                        "ORIGEN_FILA": row,
                    }
                )
    finally:
        wb.close()

    if not records:
        hoja_nombre = ws.title if ws is not None else hoja_preferida
        raise ValueError(
            f"No se detectaron registros semanales en formato horizontal en '{hoja_nombre}'."
        )

    df = pd.DataFrame(records)
    df["DIA"] = pd.to_numeric(df["DIA"], errors="coerce").astype("Int64")
    df = df[df["DIA"].notna()].copy()
    return df, ws.title if ws is not None else hoja_preferida


def _load_diccionario_matriz(path_diccionario: Path | str, cache_dir: Path | str | None) -> Dict[str, str]:
    diccionarios = cargar_diccionarios(path_diccionario, cache_dir=cache_dir)
    raw = diccionarios.get("MATRIZ", {})
    result: Dict[str, str] = {}
    for k, v in raw.items():
        nk = normalize_text(k)
        vv = safe_str(v)
        if nk and vv:
            result[nk] = vv
    return result


def _clasificar_matriz(
    semanal_df: pd.DataFrame, dic_matriz: Dict[str, str]
) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = semanal_df.copy()

    def _resolver_matriz(row: pd.Series) -> tuple[str, str]:
        actividad = safe_str(row.get("ACTIVIDAD", ""))
        tarea = safe_str(row.get("TAREA", ""))
        base = safe_str(row.get("ACTIVIDAD_DICCIONARIO", ""))
        for candidate in [actividad, tarea, base]:
            key = normalize_text(candidate)
            if key and key in dic_matriz:
                return safe_str(dic_matriz[key]), candidate
        return "", first_non_empty([actividad, tarea, base])

    resolved = df.apply(_resolver_matriz, axis=1, result_type="expand")
    resolved.columns = ["MATRIZ", "ACTIVIDAD_DICCIONARIO"]
    df["MATRIZ"] = resolved["MATRIZ"]
    df["ACTIVIDAD_DICCIONARIO"] = resolved["ACTIVIDAD_DICCIONARIO"]

    no_clasificadas = df[df["MATRIZ"].map(safe_str).eq("")].copy()
    if not no_clasificadas.empty:
        no_clasificadas = no_clasificadas[
            ["CODIGO", "ACTIVIDAD", "TAREA", "ACTIVIDAD_DICCIONARIO", "DIA", "ESTADO", "ORIGEN_HOJA", "ORIGEN_FILA"]
        ].drop_duplicates()
    else:
        no_clasificadas = pd.DataFrame(
            columns=[
                "CODIGO",
                "ACTIVIDAD",
                "TAREA",
                "ACTIVIDAD_DICCIONARIO",
                "DIA",
                "ESTADO",
                "ORIGEN_HOJA",
                "ORIGEN_FILA",
            ]
        )
    return df, no_clasificadas


def _find_header_columns(ws) -> tuple[int, int, Dict[int, int], int, int]:
    max_row, max_col = _safe_sheet_limits(ws)

    codigo_col = None
    matriz_col = None
    header_row = None
    best_header_score = -1

    codigo_priority = [normalize_column_name(v) for v in MENSUAL_CODIGO_HEADERS_PRIORITY]
    matriz_required = normalize_column_name(MENSUAL_MATRIZ_TERRENO_HEADER)

    for row in range(1, min(40, max_row) + 1):
        normalized_by_col: Dict[int, str] = {}
        for col in range(1, max_col + 1):
            normalized = normalize_column_name(ws.cell(row=row, column=col).value)
            if normalized:
                normalized_by_col[col] = normalized

        c_col = None
        for header in codigo_priority:
            for col, normalized in normalized_by_col.items():
                if normalized == header:
                    c_col = col
                    break
            if c_col is not None:
                break

        m_col = None
        for col, normalized in normalized_by_col.items():
            if normalized == matriz_required:
                m_col = col
                break

        score = int(c_col is not None) + int(m_col is not None)
        if score > best_header_score:
            best_header_score = score
            header_row = row
            codigo_col = c_col
            matriz_col = m_col

    if codigo_col is None or matriz_col is None or header_row is None:
        raise ValueError(
            "No se detectaron columnas CODIGO y MATRIZ TERRENO en programa mensual."
        )

    best_day_row = None
    best_day_count = -1
    best_day_map: Dict[int, int] = {}
    for row in range(1, min(40, max_row) + 1):
        day_to_col: Dict[int, int] = {}
        for col in range(1, max_col + 1):
            day = _extract_day(ws.cell(row=row, column=col).value)
            if day is not None:
                day_to_col[day] = col
        if len(day_to_col) > best_day_count:
            best_day_count = len(day_to_col)
            best_day_row = row
            best_day_map = day_to_col

    if best_day_row is None or not best_day_map:
        raise ValueError("No se detectaron columnas de dia (1..31) en programa mensual.")

    data_start_row = max(header_row, best_day_row) + 1
    return codigo_col, matriz_col, best_day_map, data_start_row, max_row


def _inferir_hoja_mensual_desde_semanal(semanal_df: pd.DataFrame) -> str | None:
    if "FECHA_REFERENCIA" not in semanal_df.columns:
        return None
    fechas = pd.to_datetime(semanal_df["FECHA_REFERENCIA"], errors="coerce").dropna()
    if fechas.empty:
        return None
    periodo = fechas.dt.to_period("M").mode()
    if periodo.empty:
        return None
    p = periodo.iloc[0]
    mes = MESES_ES.get(int(p.month))
    if mes is None:
        return None
    return f"{mes} {int(p.year)}"


def _parse_mes_anio_desde_titulo_hoja(sheet_title: str) -> tuple[int, int] | None:
    normalized_title = normalize_text(sheet_title)
    month_name_to_number = {normalize_text(name): month for month, name in MESES_ES.items()}

    month = None
    for token in re.split(r"[^A-Z0-9]+", normalized_title):
        if token in month_name_to_number:
            month = month_name_to_number[token]
            break

    year = None
    year_match = re.search(r"(20\d{2})", normalized_title)
    if year_match:
        year = int(year_match.group(1))

    if month is None or year is None:
        return None
    return year, month


def _build_sheet_title_from_year_month(year: int, month: int) -> str:
    mes = MESES_ES.get(month)
    if mes is None:
        raise ValueError(f"Mes invalido para construir hoja mensual: {month}")
    return f"{mes} {year}"


def _parse_mes_anio_desde_texto(text: str) -> tuple[int, int] | None:
    normalized_text = normalize_text(text)
    month_name_to_number = {normalize_text(name): month for month, name in MESES_ES.items()}

    month = None
    for token in re.split(r"[^A-Z0-9]+", normalized_text):
        if token in month_name_to_number:
            month = month_name_to_number[token]
            break

    if month is None:
        return None

    year = None
    year_match = re.search(r"(20\d{2})", normalized_text)
    if year_match:
        year = int(year_match.group(1))
    else:
        short_year_match = re.search(r"\b(\d{2})\b", normalized_text)
        if short_year_match:
            year = 2000 + int(short_year_match.group(1))

    if year is None:
        return None
    return year, month


def _filtrar_df_por_mes_anio(
    df: pd.DataFrame,
    year: int,
    month: int,
) -> pd.DataFrame:
    if df.empty or "FECHA_REFERENCIA" not in df.columns:
        return df.copy()
    fecha_series = pd.to_datetime(df["FECHA_REFERENCIA"], errors="coerce")
    years = fecha_series.dt.year
    fecha_series = fecha_series.where(years.between(2000, 2100))
    if not fecha_series.notna().any():
        return df.copy()
    mask = fecha_series.dt.year.eq(year) & fecha_series.dt.month.eq(month)
    mask = mask.fillna(False)
    return df[mask].copy()


def _filtrar_df_por_rango_fechas(
    df: pd.DataFrame,
    fecha_desde: dt.date,
    fecha_hasta: dt.date,
) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    if "FECHA_REFERENCIA" in df.columns:
        fecha_series = pd.to_datetime(df["FECHA_REFERENCIA"], errors="coerce")
        years = fecha_series.dt.year
        fecha_series = fecha_series.where(years.between(2000, 2100))
        if fecha_series.notna().any():
            desde_ts = pd.Timestamp(fecha_desde)
            hasta_ts = pd.Timestamp(fecha_hasta)
            mask = fecha_series.ge(desde_ts) & fecha_series.le(hasta_ts)
            mask = mask.fillna(False)
            return df[mask].copy()

    # Fallback para formatos semanales que solo traen numero de dia (sin fecha completa).
    if "DIA" not in df.columns:
        return df.copy()

    dia_series = pd.to_numeric(df["DIA"], errors="coerce")
    if (
        fecha_desde.year == fecha_hasta.year
        and fecha_desde.month == fecha_hasta.month
        and fecha_desde.day <= fecha_hasta.day
    ):
        mask = dia_series.between(fecha_desde.day, fecha_hasta.day, inclusive="both")
    else:
        mask = dia_series.eq(fecha_desde.day)
    mask = mask.fillna(False)
    return df[mask].copy()


def _seleccionar_hoja_mensual(
    wb,
    hoja_preferida: str | None = None,
    strict_preferida: bool = False,
):
    if hoja_preferida:
        ws_preferida = _select_sheet_by_name(wb, hoja_preferida)
        if ws_preferida is not None:
            try:
                parsed = _find_header_columns(ws_preferida)
                return ws_preferida, parsed
            except Exception as exc:
                raise ValueError(
                    f"La hoja mensual '{ws_preferida.title}' no tiene estructura valida."
                ) from exc
        if strict_preferida:
            disponibles = ", ".join(wb.sheetnames)
            raise ValueError(
                f"No se encontro la hoja mensual requerida '{hoja_preferida}'. "
                f"Hojas disponibles: {disponibles}"
            )

    best = None
    best_days = -1
    for ws in wb.worksheets:
        try:
            parsed = _find_header_columns(ws)
        except Exception:
            continue
        day_count = len(parsed[2])
        if day_count > best_days:
            best = (ws, parsed)
            best_days = day_count

    if best is None:
        raise ValueError(
            "No se detecto una hoja mensual valida con columnas CODIGO/MATRIZ y dias 1..31."
        )
    return best


def _es_estado_protegido(value: object) -> bool:
    return normalize_text(value) in ESTADOS_PROTEGIDOS


def _is_housekeeping_codigo(value: object) -> bool:
    normalized = normalize_text(value).replace(" ", "")
    return normalized in {"HOUSEKEEPING", "HOUSKEEPING"}


def _pick_housekeeping_row(candidates: List[Tuple[int, str]]) -> int | None:
    if not candidates:
        return None
    for row, matriz in candidates:
        if _canonical_matriz(matriz) == "HK":
            return row
    return candidates[0][0]


def _build_diagnostico_no_cruzados(
    no_cruzados_df: pd.DataFrame,
    matrices_por_codigo: Dict[str, set[str]],
    row_lookup: Dict[tuple[str, str], int],
    day_to_col: Dict[int, int],
) -> pd.DataFrame:
    columns = [
        "CODIGO",
        "MATRIZ_SEMANAL",
        "DIA",
        "ESTADO",
        "MOTIVO_ORIGINAL",
        "CODIGO_EXISTE_EN_MENSUAL",
        "MATRICES_EN_MENSUAL_PARA_CODIGO",
        "COINCIDE_CODIGO_MATRIZ",
        "COLUMNA_DIA_EXISTE",
    ]
    if no_cruzados_df.empty:
        return pd.DataFrame(columns=columns)

    rows: List[Dict[str, object]] = []
    for _, rec in no_cruzados_df.iterrows():
        codigo = safe_str(rec.get("CODIGO", ""))
        matriz = safe_str(rec.get("MATRIZ", ""))
        estado = safe_str(rec.get("ESTADO", ""))
        motivo = safe_str(rec.get("MOTIVO", ""))

        dia_val = rec.get("DIA")
        dia = int(dia_val) if pd.notna(dia_val) else None

        key = normalize_key(codigo, _canonical_matriz(matriz))
        matrices_mensual = sorted(matrices_por_codigo.get(codigo, set()))

        rows.append(
            {
                "CODIGO": codigo,
                "MATRIZ_SEMANAL": matriz,
                "DIA": dia,
                "ESTADO": estado,
                "MOTIVO_ORIGINAL": motivo,
                "CODIGO_EXISTE_EN_MENSUAL": codigo in matrices_por_codigo,
                "MATRICES_EN_MENSUAL_PARA_CODIGO": " | ".join(matrices_mensual),
                "COINCIDE_CODIGO_MATRIZ": key in row_lookup,
                "COLUMNA_DIA_EXISTE": (dia in day_to_col) if dia is not None else False,
            }
        )

    return pd.DataFrame(rows, columns=columns)


def actualizar_programa_mensual(
    path_programa_turno: Path | str,
    path_programa_mensual: Path | str,
    output_path_programa_mensual: Path | str,
    path_diccionario: Path | str,
    fecha_desde: dt.date | None = None,
    fecha_hasta: dt.date | None = None,
) -> pd.DataFrame:
    print("[3/5] Actualizando programa mensual...")
    output_path = Path(output_path_programa_mensual)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if fecha_desde is not None and fecha_hasta is None:
        fecha_hasta = fecha_desde
    if fecha_hasta is not None and fecha_desde is None:
        fecha_desde = fecha_hasta
    if fecha_desde is not None and fecha_hasta is not None and fecha_desde > fecha_hasta:
        raise ValueError(
            "Rango de fechas invalido: 'fecha_desde' no puede ser mayor que 'fecha_hasta'."
        )

    semanal_df, hoja_semanal_usada = _leer_semanal_horizontal(
        path_programa_turno, hoja_preferida="Control SGS"
    )
    dic_matriz = _load_diccionario_matriz(path_diccionario, cache_dir=output_path.parent)
    semanal_df, actividades_no_clasificadas_df = _clasificar_matriz(semanal_df, dic_matriz)

    work_df = semanal_df.copy()
    work_df = work_df[
        work_df["CODIGO"].map(safe_str).ne("")
        & work_df["MATRIZ"].map(safe_str).ne("")
        & work_df["ESTADO"].map(safe_str).ne("")
    ].copy()

    hoja_mensual_objetivo = None
    origen_hoja_objetivo = "auto"
    if fecha_desde is not None and fecha_hasta is not None:
        if fecha_desde.year != fecha_hasta.year or fecha_desde.month != fecha_hasta.month:
            raise ValueError(
                "El rango de fechas manual abarca mas de un mes. "
                "Para asegurar cruce correcto por hoja mensual, ejecuta una vez por mes "
                "(ej: 26/03/2026-31/03/2026 y luego 01/04/2026-01/04/2026)."
            )
        hoja_mensual_objetivo = _build_sheet_title_from_year_month(
            fecha_desde.year, fecha_desde.month
        )
        origen_hoja_objetivo = "rango_fechas"

    if hoja_mensual_objetivo is None:
        periodo_desde_nombre = _parse_mes_anio_desde_texto(Path(path_programa_mensual).stem)
        if periodo_desde_nombre is not None:
            y, m = periodo_desde_nombre
            hoja_mensual_objetivo = _build_sheet_title_from_year_month(y, m)
            origen_hoja_objetivo = "nombre_archivo"

    if hoja_mensual_objetivo is None:
        hoja_mensual_objetivo = _inferir_hoja_mensual_desde_semanal(semanal_df)
        if hoja_mensual_objetivo is not None:
            origen_hoja_objetivo = "semanal"

    wb = load_workbook(Path(path_programa_mensual))
    strict_preferida = origen_hoja_objetivo in {"rango_fechas", "nombre_archivo"}
    ws, parsed = _seleccionar_hoja_mensual(
        wb,
        hoja_preferida=hoja_mensual_objetivo,
        strict_preferida=strict_preferida,
    )
    codigo_col, matriz_col, day_to_col, data_start_row, max_row = parsed

    fuera_de_periodo = 0
    fechas_completas_disponibles = False
    if "FECHA_REFERENCIA" in work_df.columns:
        fecha_series_work = pd.to_datetime(work_df["FECHA_REFERENCIA"], errors="coerce")
        years = fecha_series_work.dt.year
        fecha_series_work = fecha_series_work.where(years.between(2000, 2100))
        fechas_completas_disponibles = fecha_series_work.notna().any()

    if fecha_desde is not None and fecha_hasta is not None:
        work_before = len(work_df)
        work_df = _filtrar_df_por_rango_fechas(work_df, fecha_desde, fecha_hasta)
        actividades_no_clasificadas_df = _filtrar_df_por_rango_fechas(
            actividades_no_clasificadas_df, fecha_desde, fecha_hasta
        )
        fuera_de_periodo = work_before - len(work_df)
    else:
        periodo_hoja = _parse_mes_anio_desde_titulo_hoja(ws.title)
        if periodo_hoja is not None:
            if not fechas_completas_disponibles:
                raise ValueError(
                    "El programa semanal no trae fechas completas (solo numero de dia). "
                    "Para evitar cruces de mes incorrectos, selecciona Fecha desde / Fecha hasta."
                )
            year, month = periodo_hoja
            work_before = len(work_df)
            work_df = _filtrar_df_por_mes_anio(work_df, year, month)
            actividades_no_clasificadas_df = _filtrar_df_por_mes_anio(
                actividades_no_clasificadas_df, year, month
            )
            fuera_de_periodo = work_before - len(work_df)

    duplicados_df = work_df[
        work_df.duplicated(subset=["CODIGO", "MATRIZ", "DIA"], keep=False)
    ].copy()
    work_df = work_df.drop_duplicates(subset=["CODIGO", "MATRIZ", "DIA"], keep="last")

    row_lookup: Dict[tuple[str, str], int] = {}
    matrices_por_codigo: Dict[str, set[str]] = {}
    rows_por_codigo: Dict[str, List[Tuple[int, str]]] = {}
    housekeeping_rows: List[Tuple[int, str]] = []
    blank_streak = 0
    for row in range(data_start_row, max_row + 1):
        codigo = safe_str(ws.cell(row=row, column=codigo_col).value)
        matriz = safe_str(ws.cell(row=row, column=matriz_col).value)
        if codigo == "" and matriz == "":
            blank_streak += 1
            if blank_streak >= 200 and row > data_start_row + 200:
                break
            continue
        blank_streak = 0
        matrices_por_codigo.setdefault(codigo, set())
        rows_por_codigo.setdefault(normalize_text(codigo), [])
        rows_por_codigo[normalize_text(codigo)].append((row, matriz))
        if matriz:
            matrices_por_codigo[codigo].add(matriz)
        if _is_housekeeping_codigo(codigo):
            housekeeping_rows.append((row, matriz))

        key = normalize_key(codigo, _canonical_matriz(matriz))
        if key == normalize_key("", ""):
            continue
        # Conserva la primera fila para evitar desfases cuando hay filas duplicadas por codigo/matriz.
        if key not in row_lookup:
            row_lookup[key] = row

    aplicados: List[Dict[str, object]] = []
    no_cruzados: List[Dict[str, object]] = []
    protegidos = 0
    no_actualizados_otro = 0

    for _, record in work_df.iterrows():
        codigo = safe_str(record["CODIGO"])
        matriz = safe_str(record["MATRIZ"])
        dia = int(record["DIA"])
        estado = safe_str(record["ESTADO"])

        key = normalize_key(codigo, _canonical_matriz(matriz))
        row = row_lookup.get(key)
        col = day_to_col.get(dia)

        if row is None and _is_housekeeping_codigo(codigo):
            direct_candidates = rows_por_codigo.get(normalize_text(codigo), [])
            row = _pick_housekeeping_row(direct_candidates)
            if row is None:
                row = _pick_housekeeping_row(housekeeping_rows)

        if row is None or col is None:
            no_cruzados.append(
                {
                    "CODIGO": codigo,
                    "MATRIZ": matriz,
                    "DIA": dia,
                    "ESTADO": estado,
                    "ACTIVIDAD": safe_str(record["ACTIVIDAD_DICCIONARIO"]),
                    "MOTIVO": "SIN_COINCIDENCIA" if row is None else "DIA_NO_ENCONTRADO",
                }
            )
            continue

        cell = ws.cell(row=row, column=col)
        valor_actual = safe_str(cell.value)
        valor_actual_norm = normalize_text(valor_actual)

        if valor_actual_norm in {"", "1"}:
            cell.value = estado
            aplicados.append(
                {
                    "CODIGO": codigo,
                    "MATRIZ": matriz,
                    "DIA": dia,
                    "ESTADO": estado,
                    "ACTIVIDAD": safe_str(record["ACTIVIDAD_DICCIONARIO"]),
                    "ORIGEN_HOJA": safe_str(record["ORIGEN_HOJA"]),
                    "ORIGEN_FILA": int(record["ORIGEN_FILA"]),
                }
            )
        elif _es_estado_protegido(valor_actual):
            protegidos += 1
        else:
            no_actualizados_otro += 1

    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise PermissionError(
            "No se pudo guardar el archivo de salida porque esta en uso: "
            f"{output_path}. Cierra el Excel abierto (y su vista previa) y vuelve a intentar."
        ) from exc

    no_cruzados_df = pd.DataFrame(
        no_cruzados,
        columns=["CODIGO", "MATRIZ", "DIA", "ESTADO", "ACTIVIDAD", "MOTIVO"],
    )
    if no_cruzados_df.empty:
        no_cruzados_df = pd.DataFrame(
            columns=["CODIGO", "MATRIZ", "DIA", "ESTADO", "ACTIVIDAD", "MOTIVO"]
        )

    diagnostico_no_cruzados_df = _build_diagnostico_no_cruzados(
        no_cruzados_df=no_cruzados_df,
        matrices_por_codigo=matrices_por_codigo,
        row_lookup=row_lookup,
        day_to_col=day_to_col,
    )

    if duplicados_df.empty:
        duplicados_df = pd.DataFrame(
            columns=[
                "CODIGO",
                "ACTIVIDAD",
                "TAREA",
                "ACTIVIDAD_DICCIONARIO",
                "DIA",
                "ESTADO",
                "MATRIZ",
                "ORIGEN_HOJA",
                "ORIGEN_FILA",
            ]
        )

    interno_dir = output_path.parent / "_interno"
    interno_dir.mkdir(parents=True, exist_ok=True)

    save_dataframe_to_excel(no_cruzados_df, interno_dir / "no_cruzados.xlsx")
    save_dataframe_to_excel(
        actividades_no_clasificadas_df,
        interno_dir / "actividades_no_clasificadas.xlsx",
    )
    save_dataframe_to_excel(
        diagnostico_no_cruzados_df,
        interno_dir / "diagnostico_no_cruzados_rapido.xlsx",
    )

    aplicados_df = pd.DataFrame(
        aplicados,
        columns=["CODIGO", "MATRIZ", "DIA", "ESTADO", "ACTIVIDAD", "ORIGEN_HOJA", "ORIGEN_FILA"],
    )
    if aplicados_df.empty:
        aplicados_df = pd.DataFrame(
            columns=["CODIGO", "MATRIZ", "DIA", "ESTADO", "ACTIVIDAD", "ORIGEN_HOJA", "ORIGEN_FILA"]
        )

    aplicados_df["PUNTO"] = aplicados_df["CODIGO"]
    aplicados_df["TIPO"] = aplicados_df["MATRIZ"]
    aplicados_df["COMENTARIO"] = aplicados_df["ACTIVIDAD"]

    print(
        "[3/5] Hojas usadas: "
        f"semanal='{hoja_semanal_usada}' | mensual='{ws.title}' "
        f"(objetivo='{hoja_mensual_objetivo or 'auto'}', origen='{origen_hoja_objetivo}')"
    )
    if fecha_desde is not None and fecha_hasta is not None:
        print(
            "[3/5] Filtro de fechas manual aplicado: "
            f"{fecha_desde.strftime('%d/%m/%Y')} -> {fecha_hasta.strftime('%d/%m/%Y')}"
        )
    print(
        "[3/5] Semanal total: "
        f"{len(semanal_df)} | clasificados matriz: {(semanal_df['MATRIZ'].map(safe_str) != '').sum()} "
        f"| fuera de periodo hoja mensual: {fuera_de_periodo} "
        f"| aplicados mensual: {len(aplicados_df)} | no cruzados: {len(no_cruzados_df)} "
        f"| actividades no clasificadas: {len(actividades_no_clasificadas_df)} "
        f"| duplicados: {len(duplicados_df)} | protegidos sin cambio: {protegidos} "
        f"| no actualizados (otro valor): {no_actualizados_otro}"
    )

    return aplicados_df[
        [
            "PUNTO",
            "TIPO",
            "DIA",
            "ESTADO",
            "COMENTARIO",
            "CODIGO",
            "MATRIZ",
            "ACTIVIDAD",
            "ORIGEN_HOJA",
            "ORIGEN_FILA",
        ]
    ]
