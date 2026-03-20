from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

from utils.excel_utils import (
    ColumnValidationError,
    read_excel_file,
    rename_columns_by_alias,
    validate_required_columns,
)
from utils.texto_utils import (
    first_non_empty,
    normalize_column_name,
    normalize_text,
    safe_str,
)

MSEWJO_ALIASES = {
    "OT": [
        "OT",
        "ORDEN_TRABAJO",
        "ORDEN_DE_TRABAJO",
        "NRO_OT",
        "N_OT",
        "NUMERO_OT",
        "NRO_ORDEN",
        "NO_ORDEN",
    ],
    "FECHA": [
        "FECHA",
        "FECHA_DE_EJECUCION",
        "FECHA_MUESTREO",
        "FECHA_PROGRAMADA",
        "FECHA_EJECUCION",
        "FEC_MUESTRA",
        "FEC",
    ],
    "PUNTO": ["PUNTO", "PUNTO_MONITOREO", "ESTACION", "PTO"],
    "TIPO": ["TIPO", "MATRIZ", "TIPO_MATRIZ"],
    "DESCRIPCION": ["DESCRIPCION", "DETALLE", "OBSERVACION", "OBSERVACIONES"],
    "DESCRIPCION_TAREA_2": [
        "DESCRIPCION_DE_TAREA_PROGRAMADA_2",
        "DESCRIPCION_TAREA_PROGRAMADA_2",
        "DESC_TAREA_PROGRAMADA_2",
        "HS",
    ],
    "GRUPO_TRAB": ["GRUPO_TRAB", "GRUPO_DE_TRABAJO", "CJ"],
}

TIPOS_CONOCIDOS = [
    "AGUA",
    "AIRE",
    "RUIDO",
    "SUELO",
    "EMISION",
    "EFLUENTE",
]

MSEWJO_FIXED_COLUMNS = {
    "OT": "D",
    "DESCRIPCION": "E",
    "DESCRIPCION_TAREA_2": "HS",
    "FECHA_INICIO_PLAN": "BD",
    "FECHA": "BF",
    "GRUPO_TRAB": "CJ",
}


def _match_alias(column_name: object, alias: object) -> bool:
    column_norm = normalize_column_name(column_name)
    alias_norm = normalize_column_name(alias)
    if not column_norm or not alias_norm:
        return False
    if column_norm == alias_norm:
        return True
    padded = f"_{column_norm}_"
    return f"_{alias_norm}_" in padded


def _detectar_fila_encabezado(path_msewjo: str | Path, max_rows: int = 30) -> int | None:
    preview = pd.read_excel(Path(path_msewjo), header=None)
    if preview.empty:
        return None

    required = ["OT", "FECHA"]
    aliases_norm = {
        canonical: [normalize_column_name(alias) for alias in MSEWJO_ALIASES[canonical]]
        for canonical in required
    }

    best_row = None
    best_score = -1

    max_scan = min(max_rows, len(preview))
    for row_idx in range(max_scan):
        values = preview.iloc[row_idx].tolist()
        score = 0
        for canonical in required:
            if any(
                _match_alias(value, alias)
                for value in values
                for alias in aliases_norm[canonical]
            ):
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx

    if best_score < len(required):
        return None
    return best_row


def _leer_msewjo(path_msewjo: str | Path) -> pd.DataFrame:
    df = read_excel_file(path_msewjo)
    df = rename_columns_by_alias(df, MSEWJO_ALIASES)
    if all(c in df.columns for c in ["OT", "FECHA"]):
        return df

    header_row = _detectar_fila_encabezado(path_msewjo)
    if header_row is not None and header_row != 0:
        df_alt = pd.read_excel(Path(path_msewjo), header=header_row)
        df_alt = rename_columns_by_alias(df_alt, MSEWJO_ALIASES)
        if all(c in df_alt.columns for c in ["OT", "FECHA"]):
            return df_alt

    df_fixed = _leer_msewjo_por_columnas_fijas(path_msewjo)
    if all(c in df_fixed.columns for c in ["OT", "FECHA"]) and not df_fixed.empty:
        return df_fixed

    return df


def _leer_msewjo_por_columnas_fijas(path_msewjo: str | Path) -> pd.DataFrame:
    """
    Fallback inspirado en el notebook de referencia:
    D=OT, E=Descripcion, HS=Descripcion tarea 2, BF=Fecha fin plan, BD=Fecha inicio.
    """
    wb = load_workbook(Path(path_msewjo), data_only=True, read_only=True)
    try:
        ws = wb.active

        col_idx = {
            key: column_index_from_string(col)
            for key, col in MSEWJO_FIXED_COLUMNS.items()
        }
        min_col = min(col_idx.values())
        max_col = max(col_idx.values())

        def _value_from_row(row_cells, absolute_col: int):
            offset = absolute_col - min_col
            if offset < 0 or offset >= len(row_cells):
                return None
            return row_cells[offset].value

        rows = []
        for row_cells in ws.iter_rows(min_row=2, min_col=min_col, max_col=max_col):
            ot = safe_str(_value_from_row(row_cells, col_idx["OT"]))
            if not ot:
                continue

            desc_1 = safe_str(_value_from_row(row_cells, col_idx["DESCRIPCION"]))
            desc_2 = safe_str(_value_from_row(row_cells, col_idx["DESCRIPCION_TAREA_2"]))
            descripcion = " ".join(part for part in [desc_1, desc_2] if part).strip()

            fecha = _value_from_row(row_cells, col_idx["FECHA"])
            if fecha in (None, ""):
                fecha = _value_from_row(row_cells, col_idx["FECHA_INICIO_PLAN"])

            rows.append(
                {
                    "OT": ot,
                    "FECHA": fecha,
                    "DESCRIPCION": descripcion,
                    "DESCRIPCION_TAREA_2": desc_2,
                    "GRUPO_TRAB": safe_str(_value_from_row(row_cells, col_idx["GRUPO_TRAB"])),
                }
            )

        return pd.DataFrame(rows)
    finally:
        wb.close()


def _extraer_punto(texto: object) -> str:
    text = normalize_text(texto)
    if not text:
        return ""

    patterns = [
        r"(?:PUNTO|ESTACION|PTO)\s*[:\-]\s*([A-Z0-9/_\-\s]+)",
        r"(?:PUNTO|ESTACION|PTO)\s+([A-Z0-9/_\-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).strip()
    return ""


def _extraer_tipo(texto: object) -> str:
    text = normalize_text(texto)
    for tipo in TIPOS_CONOCIDOS:
        if tipo in text:
            return tipo
    return ""


def limpiar_msewjo(path_msewjo: str | Path) -> pd.DataFrame:
    print("[1/5] Leyendo MSEWJO...")
    df = _leer_msewjo(path_msewjo)
    try:
        validate_required_columns(df, ["OT", "FECHA"], "MSEWJO")
    except ColumnValidationError as exc:
        cols = ", ".join(str(c) for c in df.columns[:30])
        raise ColumnValidationError(
            f"{exc}. Columnas detectadas: {cols or '(sin columnas)'}"
        ) from exc

    if "DESCRIPCION" not in df.columns:
        df["DESCRIPCION"] = ""
    if "DESCRIPCION_TAREA_2" not in df.columns:
        df["DESCRIPCION_TAREA_2"] = ""
    if "GRUPO_TRAB" not in df.columns:
        df["GRUPO_TRAB"] = ""

    if "PUNTO" not in df.columns:
        df["PUNTO"] = ""
    if "TIPO" not in df.columns:
        df["TIPO"] = ""

    df["OT"] = df["OT"].map(safe_str)
    df["DESCRIPCION_TAREA_2"] = df["DESCRIPCION_TAREA_2"].map(safe_str)
    df["GRUPO_TRAB"] = df["GRUPO_TRAB"].map(safe_str)
    df["PUNTO"] = df.apply(
        lambda r: first_non_empty(
            [
                r.get("PUNTO"),
                r.get("DESCRIPCION_TAREA_2"),
                _extraer_punto(r.get("DESCRIPCION")),
            ]
        ),
        axis=1,
    )
    df["TIPO"] = df.apply(
        lambda r: first_non_empty(
            [
                r.get("TIPO"),
                r.get("GRUPO_TRAB"),
                _extraer_tipo(r.get("DESCRIPCION")),
            ]
        ),
        axis=1,
    )

    print("[1/5] Convirtiendo fechas...")
    df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce", dayfirst=True)
    before_dates = len(df)
    df = df[df["FECHA"].notna()].copy()
    dropped_dates = before_dates - len(df)
    if dropped_dates:
        print(f"[1/5] Filas descartadas por fecha invalida: {dropped_dates}")

    df["DIA"] = df["FECHA"].dt.day.astype("Int64")

    print("[1/5] Eliminando duplicados...")
    dedup_subset = [c for c in ["OT", "PUNTO", "TIPO", "FECHA"] if c in df.columns]
    before_dedup = len(df)
    df = df.drop_duplicates(subset=dedup_subset).copy()
    print(f"[1/5] Duplicados eliminados: {before_dedup - len(df)}")

    df["PUNTO"] = df["PUNTO"].map(safe_str)
    df["TIPO"] = df["TIPO"].map(safe_str)
    df["DESCRIPCION"] = df["DESCRIPCION"].map(safe_str)

    # Mantiene columnas clave al inicio para facilitar trazabilidad.
    key_cols = ["OT", "PUNTO", "TIPO", "FECHA", "DIA", "DESCRIPCION"]
    other_cols = [c for c in df.columns if c not in key_cols]
    ordered_cols = key_cols + other_cols
    return df[ordered_cols]
