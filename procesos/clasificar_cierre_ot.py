from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import load_workbook

from utils.excel_utils import (
    save_dataframe_to_excel,
    validate_required_columns,
)
from utils.texto_utils import normalize_column_name, normalize_text, safe_str

CACHE_FILENAME = "diccionario_cache.json"
META_FILENAME = "diccionario_meta.json"

MATRIZ_COL_ALIASES = {
    "ACTIVIDAD": [
        "ACTIVIDAD",
        "ACTIVIDAD_TAREA",
        "ACTIVIDAD O TAREA",
        "TAREA",
        "DESCRIPCION",
        "DESCRIPCION_TAREA",
    ],
    "MATRIZ_TERRENO": [
        "MATRIZ_TERRENO",
        "MATRIZ TERRENO",
        "MATRIZ",
        "TIPO_MATRIZ",
    ],
}

CIERRE_COL_ALIASES = {
    "ESTADO": ["ESTADO", "STATUS", "RESULTADO"],
    "PALABRA_CLAVE_COMENTARIO": [
        "PALABRA_CLAVE_COMENTARIO",
        "PALABRA CLAVE COMENTARIO",
        "PALABRA_CLAVE",
        "KEYWORD",
    ],
    "CODIGO_CIERRE": [
        "CODIGO_CIERRE",
        "CODIGO_DE_CIERRE",
        "CODIGO",
        "CÓDIGO_DE_CIERRE",
    ],
}

ACTIVIDAD_INPUT_ALIASES = [
    "ACTIVIDAD",
    "ACTIVIDAD_TAREA",
    "DESCRIPCION_TAREA_2",
    "DESCRIPCION",
    "TAREA",
    "TIPO",
]


def limpiar_texto(value: object) -> str:
    return normalize_text(value)


def _cache_paths(path_diccionario_excel: Path | str, cache_dir: Path | str | None = None) -> tuple[Path, Path]:
    base = Path(cache_dir) if cache_dir is not None else Path(path_diccionario_excel).parent
    base.mkdir(parents=True, exist_ok=True)
    return base / CACHE_FILENAME, base / META_FILENAME


def _read_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, dict):
        raise ValueError(f"JSON invalido en {path}")
    return data


def _write_json(path: Path, payload: Dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)


def _find_column(df: pd.DataFrame, aliases: Iterable[str]) -> str | None:
    norm_map = {normalize_column_name(col): col for col in df.columns}
    for alias in aliases:
        candidate = norm_map.get(normalize_column_name(alias))
        if candidate is not None:
            return candidate
    return None


def _sheet_matches(df: pd.DataFrame, aliases_map: Dict[str, Iterable[str]]) -> bool:
    for alias_group in aliases_map.values():
        if _find_column(df, alias_group) is None:
            return False
    return True


def _load_sheets(path_diccionario_excel: Path | str) -> Dict[str, pd.DataFrame]:
    excel = pd.ExcelFile(Path(path_diccionario_excel))
    return {sheet_name: excel.parse(sheet_name=sheet_name) for sheet_name in excel.sheet_names}


def _select_sheet(
    sheets: Dict[str, pd.DataFrame],
    aliases_map: Dict[str, Iterable[str]],
    preferred_names: Iterable[str],
) -> tuple[str | None, pd.DataFrame | None]:
    preferred_norm = [normalize_column_name(name) for name in preferred_names]
    for sheet_name, df in sheets.items():
        if normalize_column_name(sheet_name) in preferred_norm and _sheet_matches(df, aliases_map):
            return sheet_name, df

    for sheet_name, df in sheets.items():
        if _sheet_matches(df, aliases_map):
            return sheet_name, df
    return None, None


def _build_diccionarios_from_excel(path_diccionario_excel: Path | str) -> Dict[str, Any]:
    sheets = _load_sheets(path_diccionario_excel)

    _, matriz_df = _select_sheet(
        sheets=sheets,
        aliases_map=MATRIZ_COL_ALIASES,
        preferred_names=["MATRIZ_TERRENO", "MATRIZ TERRENO"],
    )
    _, cierre_df = _select_sheet(
        sheets=sheets,
        aliases_map=CIERRE_COL_ALIASES,
        preferred_names=["COD_CIERRE", "COD CIERRE", "CLASIFICACION", "MATRIZ_CLASIFICACION"],
    )

    matriz_dict: Dict[str, str] = {}
    if matriz_df is not None:
        actividad_col = _find_column(matriz_df, MATRIZ_COL_ALIASES["ACTIVIDAD"])
        matriz_col = _find_column(matriz_df, MATRIZ_COL_ALIASES["MATRIZ_TERRENO"])
        if actividad_col and matriz_col:
            for _, row in matriz_df.iterrows():
                actividad_norm = limpiar_texto(row.get(actividad_col))
                matriz_val = safe_str(row.get(matriz_col))
                if actividad_norm and matriz_val:
                    matriz_dict[actividad_norm] = matriz_val

    cierre_dict: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    if cierre_df is not None:
        estado_col = _find_column(cierre_df, CIERRE_COL_ALIASES["ESTADO"])
        keyword_col = _find_column(cierre_df, CIERRE_COL_ALIASES["PALABRA_CLAVE_COMENTARIO"])
        codigo_col = _find_column(cierre_df, CIERRE_COL_ALIASES["CODIGO_CIERRE"])
        if estado_col and keyword_col and codigo_col:
            for _, row in cierre_df.iterrows():
                estado_norm = limpiar_texto(row.get(estado_col))
                keyword_norm = limpiar_texto(row.get(keyword_col))
                codigo = safe_str(row.get(codigo_col))
                if estado_norm and codigo:
                    cierre_dict[estado_norm].append({"keyword": keyword_norm, "codigo": codigo})

    return {"MATRIZ": matriz_dict, "CIERRE": dict(cierre_dict)}


def cargar_diccionarios(
    path_diccionario_excel: Path | str,
    cache_dir: Path | str | None = None,
) -> Dict[str, Any]:
    """
    Lee diccionarios desde Excel y usa cache JSON cuando no hay cambios.
    Regenera cache automaticamente si cambia el Excel fuente.
    """
    source_path = Path(path_diccionario_excel)
    if not source_path.exists():
        raise FileNotFoundError(f"No existe el archivo de diccionarios: {source_path}")

    cache_warning = ""
    try:
        cache_path, meta_path = _cache_paths(source_path, cache_dir=cache_dir)
    except Exception as exc:
        fallback = Path.cwd() / "_diccionario_cache"
        cache_path, meta_path = _cache_paths(source_path, cache_dir=fallback)
        cache_warning = (
            "[ADVERTENCIA] No se pudo usar carpeta de cache junto al diccionario; "
            f"se usara fallback en {fallback}. Motivo: {exc}"
        )

    if cache_warning:
        print(cache_warning)
    source_mtime = os.path.getmtime(source_path)

    if cache_path.exists() and meta_path.exists():
        try:
            meta = _read_json(meta_path)
            cache = _read_json(cache_path)
            cached_mtime = float(meta.get("source_mtime", -1))
            if cached_mtime == float(source_mtime):
                if "MATRIZ" in cache and "CIERRE" in cache:
                    print("[4/5] Diccionarios cargados desde cache.")
                    return cache
        except Exception:
            pass

    print("[4/5] Leyendo diccionarios desde Excel y regenerando cache...")
    diccionarios = _build_diccionarios_from_excel(source_path)
    try:
        _write_json(cache_path, diccionarios)
        _write_json(
            meta_path,
            {
                "source_path": str(source_path.resolve()),
                "source_mtime": float(source_mtime),
            },
        )
    except Exception as exc:
        print(
            "[ADVERTENCIA] No se pudo escribir cache de diccionarios. "
            f"Se continuara sin cache. Motivo: {exc}"
        )
    return diccionarios


def clasificar_matriz(
    actividad_tarea: object,
    diccionario_matriz: Dict[str, str],
    nuevas_actividades: List[str],
) -> str:
    actividad_original = safe_str(actividad_tarea)
    actividad_norm = limpiar_texto(actividad_tarea)
    if not actividad_norm:
        return ""

    matriz = safe_str(diccionario_matriz.get(actividad_norm, ""))
    if not matriz:
        nuevas_actividades.append(actividad_original)
    return matriz


def clasificar_cierre(
    estado: object,
    comentario: object,
    diccionario_cierre: Dict[str, List[Dict[str, str]]],
    comentarios_no_clasificados: List[Dict[str, str]],
) -> str:
    estado_norm = limpiar_texto(estado)
    comentario_norm = limpiar_texto(comentario)
    reglas_estado = diccionario_cierre.get(estado_norm, [])

    codigo = ""
    fallback = ""
    for rule in reglas_estado:
        keyword_norm = limpiar_texto(rule.get("keyword", ""))
        candidate_code = safe_str(rule.get("codigo", ""))
        if not candidate_code:
            continue

        if keyword_norm:
            if keyword_norm in comentario_norm:
                codigo = candidate_code
                break
        elif not fallback:
            fallback = candidate_code

    if not codigo:
        codigo = fallback

    if not codigo:
        comentario_original = safe_str(comentario)
        if comentario_original:
            comentarios_no_clasificados.append(
                {
                    "ESTADO": safe_str(estado),
                    "COMENTARIO": comentario_original,
                }
            )
    return codigo


def _dedup_strings(values: Iterable[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for raw in values:
        value = safe_str(raw)
        if not value:
            continue
        key = limpiar_texto(value)
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _dedup_comment_rows(rows: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    result: List[Dict[str, str]] = []
    for row in rows:
        estado = safe_str(row.get("ESTADO"))
        comentario = safe_str(row.get("COMENTARIO"))
        if not comentario:
            continue
        key = (limpiar_texto(estado), limpiar_texto(comentario))
        if key in seen:
            continue
        seen.add(key)
        result.append({"ESTADO": estado, "COMENTARIO": comentario})
    return result


def detectar_nuevos_valores(
    nuevas_actividades: List[str],
    comentarios_no_clasificados: List[Dict[str, str]],
    modo: str = "automatico",
    export_dir: Path | str | None = None,
    diccionarios: Dict[str, Any] | None = None,
) -> Dict[str, List[Any]]:
    modo_norm = limpiar_texto(modo).lower()
    nuevas_actividades_u = _dedup_strings(nuevas_actividades)
    comentarios_no_clasificados_u = _dedup_comment_rows(comentarios_no_clasificados)

    if not nuevas_actividades_u and not comentarios_no_clasificados_u:
        return {
            "nuevas_actividades": [],
            "comentarios_no_clasificados": [],
        }

    if modo_norm in {"automatico", "auto"}:
        if nuevas_actividades_u:
            print("[ADVERTENCIA] Nuevas actividades sin matriz detectadas:")
            for item in nuevas_actividades_u:
                print(f"  - {item}")
        if comentarios_no_clasificados_u:
            print("[ADVERTENCIA] Comentarios sin clasificar detectados:")
            for item in comentarios_no_clasificados_u:
                print(f"  - ESTADO={item['ESTADO']} | COMENTARIO={item['COMENTARIO']}")

    elif modo_norm in {"interactivo", "interactiva"}:
        if diccionarios is None:
            diccionarios = {"MATRIZ": {}, "CIERRE": {}}
        matriz_dict = diccionarios.setdefault("MATRIZ", {})
        cierre_dict = diccionarios.setdefault("CIERRE", {})

        for actividad in nuevas_actividades_u:
            respuesta = input(
                f"[INTERACTIVO] Matriz para actividad '{actividad}' (enter para omitir): "
            ).strip()
            if respuesta:
                matriz_dict[limpiar_texto(actividad)] = respuesta

        for row in comentarios_no_clasificados_u:
            estado = row["ESTADO"]
            comentario = row["COMENTARIO"]
            respuesta = input(
                f"[INTERACTIVO] Codigo cierre para ESTADO='{estado}' y comentario "
                f"'{comentario}' (enter para omitir): "
            ).strip()
            if respuesta:
                estado_norm = limpiar_texto(estado)
                cierre_dict.setdefault(estado_norm, []).append(
                    {"keyword": limpiar_texto(comentario), "codigo": respuesta}
                )

    elif modo_norm in {"controlado", "exportar"}:
        export_base = Path(export_dir) if export_dir is not None else Path.cwd()
        export_base.mkdir(parents=True, exist_ok=True)

        if nuevas_actividades_u:
            df_new_acts = pd.DataFrame({"ACTIVIDAD": nuevas_actividades_u})
            save_dataframe_to_excel(df_new_acts, export_base / "nuevas_actividades.xlsx")
            print(
                "[ADVERTENCIA] Se exportaron actividades nuevas a "
                f"{export_base / 'nuevas_actividades.xlsx'}"
            )

        if comentarios_no_clasificados_u:
            df_new_comments = pd.DataFrame(comentarios_no_clasificados_u)
            save_dataframe_to_excel(
                df_new_comments, export_base / "comentarios_no_clasificados.xlsx"
            )
            print(
                "[ADVERTENCIA] Se exportaron comentarios no clasificados a "
                f"{export_base / 'comentarios_no_clasificados.xlsx'}"
            )

    else:
        print(
            f"[ADVERTENCIA] Modo de nuevos valores no reconocido: '{modo}'. "
            "Se usara modo automatico."
        )
        return detectar_nuevos_valores(
            nuevas_actividades=nuevas_actividades_u,
            comentarios_no_clasificados=comentarios_no_clasificados_u,
            modo="automatico",
            export_dir=export_dir,
            diccionarios=diccionarios,
        )

    return {
        "nuevas_actividades": nuevas_actividades_u,
        "comentarios_no_clasificados": comentarios_no_clasificados_u,
    }


def _actualizar_diccionario_fuente(
    path_diccionario_excel: Path | str,
    nuevas_actividades: List[str],
    comentarios_no_clasificados: List[Dict[str, str]],
) -> Dict[str, int]:
    """
    Actualiza el Excel de diccionario agregando filas nuevas pendientes:
    - MATRIZ_TERRENO: actividad nueva con MATRIZ TERRENO en blanco.
    - COD_CIERRE: estado/comentario nuevo con CODIGO en blanco.
    """
    source = Path(path_diccionario_excel)
    if not source.exists():
        raise FileNotFoundError(f"No existe el archivo de diccionario para actualizar: {source}")

    sheets = _load_sheets(source)
    matriz_sheet_name, matriz_df = _select_sheet(
        sheets=sheets,
        aliases_map=MATRIZ_COL_ALIASES,
        preferred_names=["MATRIZ_TERRENO", "MATRIZ TERRENO"],
    )
    cierre_sheet_name, cierre_df = _select_sheet(
        sheets=sheets,
        aliases_map=CIERRE_COL_ALIASES,
        preferred_names=["COD_CIERRE", "COD CIERRE", "CLASIFICACION", "MATRIZ_CLASIFICACION"],
    )

    matriz_sheet_name = matriz_sheet_name or "MATRIZ_TERRENO"
    cierre_sheet_name = cierre_sheet_name or "COD_CIERRE"

    if matriz_df is None:
        matriz_df = pd.DataFrame(columns=["ACTIVIDAD", "MATRIZ TERRENO"])
    if cierre_df is None:
        cierre_df = pd.DataFrame(
            columns=["ESTADO", "PALABRA CLAVE COMENTARIO", "CÓDIGO DE CIERRE"]
        )

    actividad_col = _find_column(matriz_df, MATRIZ_COL_ALIASES["ACTIVIDAD"]) or "ACTIVIDAD"
    matriz_col = _find_column(matriz_df, MATRIZ_COL_ALIASES["MATRIZ_TERRENO"]) or "MATRIZ TERRENO"
    estado_col = _find_column(cierre_df, CIERRE_COL_ALIASES["ESTADO"]) or "ESTADO"
    keyword_col = _find_column(cierre_df, CIERRE_COL_ALIASES["PALABRA_CLAVE_COMENTARIO"]) or "PALABRA CLAVE COMENTARIO"
    codigo_col = _find_column(cierre_df, CIERRE_COL_ALIASES["CODIGO_CIERRE"]) or "CÓDIGO DE CIERRE"

    for col in [actividad_col, matriz_col]:
        if col not in matriz_df.columns:
            matriz_df[col] = ""
    for col in [estado_col, keyword_col, codigo_col]:
        if col not in cierre_df.columns:
            cierre_df[col] = ""

    existing_acts = {limpiar_texto(v) for v in matriz_df[actividad_col].tolist() if safe_str(v)}
    existing_pairs = {
        (limpiar_texto(r.get(estado_col)), limpiar_texto(r.get(keyword_col)))
        for _, r in cierre_df.iterrows()
        if safe_str(r.get(estado_col)) or safe_str(r.get(keyword_col))
    }

    add_acts = 0
    add_comments = 0

    for actividad in _dedup_strings(nuevas_actividades):
        key = limpiar_texto(actividad)
        if not key or key in existing_acts:
            continue
        matriz_df = pd.concat(
            [
                matriz_df,
                pd.DataFrame([{actividad_col: actividad, matriz_col: ""}]),
            ],
            ignore_index=True,
        )
        existing_acts.add(key)
        add_acts += 1

    for item in _dedup_comment_rows(comentarios_no_clasificados):
        estado = safe_str(item.get("ESTADO"))
        keyword = safe_str(item.get("COMENTARIO"))
        pair = (limpiar_texto(estado), limpiar_texto(keyword))
        if pair in existing_pairs or not (pair[0] or pair[1]):
            continue
        cierre_df = pd.concat(
            [
                cierre_df,
                pd.DataFrame(
                    [{estado_col: estado, keyword_col: keyword, codigo_col: ""}]
                ),
            ],
            ignore_index=True,
        )
        existing_pairs.add(pair)
        add_comments += 1

    if add_acts == 0 and add_comments == 0:
        return {"actividades_agregadas": 0, "comentarios_agregados": 0}

    sheets[matriz_sheet_name] = matriz_df
    sheets[cierre_sheet_name] = cierre_df

    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    return {
        "actividades_agregadas": add_acts,
        "comentarios_agregados": add_comments,
    }


def _detectar_columna_actividad(registros_df: pd.DataFrame) -> str | None:
    norm_map = {normalize_column_name(col): col for col in registros_df.columns}
    for alias in ACTIVIDAD_INPUT_ALIASES:
        detected = norm_map.get(normalize_column_name(alias))
        if detected is not None:
            return detected
    return None


def clasificar_registros(
    registros_df: pd.DataFrame,
    path_matriz_clasificacion: Path | str,
    modo_nuevos_valores: str = "automatico",
    export_dir: Path | str | None = None,
    actualizar_diccionario_fuente: bool = False,
) -> pd.DataFrame:
    print("[4/5] Clasificando registros...")
    validate_required_columns(
        registros_df, ["PUNTO", "TIPO", "DIA", "ESTADO", "COMENTARIO"], "registros turno"
    )

    try:
        diccionarios = cargar_diccionarios(
            path_matriz_clasificacion,
            cache_dir=export_dir,
        )
    except Exception as exc:
        raise RuntimeError(
            "No fue posible cargar diccionarios. Verifica el archivo Excel de diccionarios."
        ) from exc

    dic_matriz: Dict[str, str] = diccionarios.get("MATRIZ", {})
    dic_cierre: Dict[str, List[Dict[str, str]]] = diccionarios.get("CIERRE", {})

    activity_col = _detectar_columna_actividad(registros_df)
    if activity_col is None:
        print(
            "[ADVERTENCIA] No se detecto columna de actividad/tarea en registros. "
            "Se intentara con columna TIPO."
        )
        activity_col = "TIPO"

    classified = registros_df.copy()
    codigos: List[str] = []
    matrices: List[str] = []

    nuevas_actividades: List[str] = []
    comentarios_no_clasificados: List[Dict[str, str]] = []

    for _, row in classified.iterrows():
        actividad = row.get(activity_col, "")
        estado = row.get("ESTADO", "")
        comentario = row.get("COMENTARIO", "")

        matriz = clasificar_matriz(
            actividad_tarea=actividad,
            diccionario_matriz=dic_matriz,
            nuevas_actividades=nuevas_actividades,
        )
        codigo = clasificar_cierre(
            estado=estado,
            comentario=comentario,
            diccionario_cierre=dic_cierre,
            comentarios_no_clasificados=comentarios_no_clasificados,
        )

        matrices.append(matriz)
        codigos.append(codigo)

    classified["MATRIZ_TERRENO"] = matrices
    classified["CODIGO_CIERRE"] = codigos

    novedades = detectar_nuevos_valores(
        nuevas_actividades=nuevas_actividades,
        comentarios_no_clasificados=comentarios_no_clasificados,
        modo=modo_nuevos_valores,
        export_dir=export_dir,
        diccionarios=diccionarios,
    )

    if actualizar_diccionario_fuente and (
        novedades["nuevas_actividades"] or novedades["comentarios_no_clasificados"]
    ):
        try:
            resumen_update = _actualizar_diccionario_fuente(
                path_diccionario_excel=path_matriz_clasificacion,
                nuevas_actividades=novedades["nuevas_actividades"],
                comentarios_no_clasificados=novedades["comentarios_no_clasificados"],
            )
            if resumen_update["actividades_agregadas"] or resumen_update["comentarios_agregados"]:
                print(
                    "[4/5] Diccionario fuente actualizado: "
                    f"{resumen_update['actividades_agregadas']} actividades nuevas y "
                    f"{resumen_update['comentarios_agregados']} comentarios clave nuevos."
                )
        except Exception as exc:
            print(
                "[ADVERTENCIA] No se pudo actualizar el diccionario fuente: "
                f"{exc}"
            )

    actividades_contenido = classified[activity_col].map(safe_str).str.strip() != ""
    comentarios_contenido = classified["COMENTARIO"].map(safe_str).str.strip() != ""
    missing_matriz = int((actividades_contenido & (classified["MATRIZ_TERRENO"].map(safe_str) == "")).sum())
    missing_cierre = int((comentarios_contenido & (classified["CODIGO_CIERRE"].map(safe_str) == "")).sum())

    if missing_matriz:
        print(
            "[ADVERTENCIA] Hay actividades sin matriz asignada: "
            f"{missing_matriz}. Revisa nuevas actividades detectadas."
        )
    if missing_cierre:
        print(
            "[ADVERTENCIA] Hay comentarios sin codigo de cierre: "
            f"{missing_cierre}. Revisa comentarios no clasificados."
        )

    print(
        f"[4/5] Registros clasificados con codigo: "
        f"{(classified['CODIGO_CIERRE'].astype(str).str.strip() != '').sum()} "
        f"de {len(classified)}"
    )
    if novedades["nuevas_actividades"] or novedades["comentarios_no_clasificados"]:
        print(
            "[4/5] Novedades detectadas: "
            f"{len(novedades['nuevas_actividades'])} actividades nuevas, "
            f"{len(novedades['comentarios_no_clasificados'])} comentarios no clasificados."
        )
    return classified


# =========================
# Etapa 3 (final) - Cierre OT
# =========================

GRUPOS_TRABAJO_ETAPA3 = {"SIGVA", "SIGVANC"}
ESTADO_PRIORIDAD_ETAPA3 = {"": 0, "R": 1, "RR": 2, "NR": 3, "S": 4}
MESES_ES_ETAPA3 = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

ALIAS_COL_CONECTOR = ["CONECTOR"]
ALIAS_COL_COMENTARIO_MENSUAL = [
    "OBS TERRENO",
    "OBS_TERRENO",
    "OBSERVACION",
    "COMENTARIO",
    "NOTA",
]

ALIAS_COL_DICCIONARIO_PALABRA = [
    "PALABRA CLAVE COMENTARIO",
    "PALABRA_CLAVE_COMENTARIO",
    "PALABRA CLAVE",
    "PALABRA_CLAVE",
    "KEYWORD",
]
ALIAS_COL_DICCIONARIO_CODIGO = [
    "CODIGO DE CIERRE",
    "CÓDIGO DE CIERRE",
    "CODIGO_CIERRE",
    "CODIGO DE CIERRE",
    "CODIGO",
]

ALIAS_COL_CIERRE_DESC2 = [
    "Descripción de tarea programada 2",
    "Descripcion de tarea programada 2",
]
ALIAS_COL_CIERRE_MATRIZ = ["Matriz Terreno", "MATRIZ TERRENO"]
ALIAS_COL_CIERRE_GRUPO = ["Grupo trab", "GRUPO TRAB"]
ALIAS_COL_CIERRE_CODIGO = ["CODIGO DE CIERRE", "CODIGO_CIERRE"]
ALIAS_COL_CIERRE_COMENTARIO_ADICIONA = ["Comentario adiciona", "COMENTARIO ADICIONA"]
ALIAS_COL_CIERRE_FECHA_INICIO = ["Fecha de inicio del plan", "FECHA DE INICIO DEL PLAN"]
ALIAS_COL_CIERRE_FECHA_FIN = [
    "Fecha de finalización planificada",
    "Fecha de finalizacion planificada",
    "FECHA DE FINALIZACION PLANIFICADA",
]


def _normalizar_conector(value: object) -> str:
    text = normalize_text(value)
    text = text.replace(" ", "")
    text = text.replace("\\", "/")
    while "//" in text:
        text = text.replace("//", "/")
    return text


def _normalizar_llave(desc_tarea_2: object, matriz_terreno: object) -> str:
    return _normalizar_conector(f"{safe_str(desc_tarea_2)}/{safe_str(matriz_terreno)}")


def _extraer_dia(value: object) -> int | None:
    if isinstance(value, int):
        return value if 1 <= value <= 31 else None
    if isinstance(value, float) and value.is_integer():
        day = int(value)
        return day if 1 <= day <= 31 else None
    if hasattr(value, "day"):
        try:
            day = int(value.day)
            return day if 1 <= day <= 31 else None
        except Exception:
            return None
    text = safe_str(value)
    if text.isdigit():
        day = int(text)
        return day if 1 <= day <= 31 else None
    return None


def _dia_desde_nombre_hoja(title: object) -> int | None:
    match = re.search(r"\b([0-9]{1,2})\b", safe_str(title))
    if not match:
        return None
    day = int(match.group(1))
    return day if 1 <= day <= 31 else None


def _to_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = safe_str(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            continue
    return None


def _iter_date_range(fecha_inicio: date | None, fecha_fin: date | None) -> List[date]:
    if fecha_inicio is None and fecha_fin is None:
        return []
    if fecha_inicio is None:
        return [fecha_fin] if fecha_fin is not None else []
    if fecha_fin is None:
        return [fecha_inicio]
    start = min(fecha_inicio, fecha_fin)
    end = max(fecha_inicio, fecha_fin)
    total_days = (end - start).days + 1
    if total_days <= 0:
        return []
    # Limita rangos anormales para evitar cruces involuntarios muy amplios.
    total_days = min(total_days, 62)
    return [start + timedelta(days=i) for i in range(total_days)]


def _periodo_desde_titulo_hoja_mensual(title: object) -> tuple[int, int] | None:
    text = normalize_text(title)
    month = None
    for token in re.split(r"[^A-Z0-9]+", text):
        if token in MESES_ES_ETAPA3:
            month = MESES_ES_ETAPA3[token]
            break
    if month is None:
        return None
    year_match = re.search(r"(20[0-9]{2})", text)
    if not year_match:
        return None
    return int(year_match.group(1)), month


def _buscar_columna_por_aliases_ws(ws, header_row: int, aliases: Iterable[str]) -> int | None:
    aliases_norm = {normalize_column_name(alias) for alias in aliases}
    for col in range(1, ws.max_column + 1):
        header = normalize_column_name(ws.cell(row=header_row, column=col).value)
        if header in aliases_norm:
            return col
    return None


def _mapa_dias_en_fila(ws, row: int) -> Dict[int, int]:
    day_map: Dict[int, int] = {}
    if row < 1:
        return day_map
    for col in range(1, ws.max_column + 1):
        day = _extraer_dia(ws.cell(row=row, column=col).value)
        if day is not None and day not in day_map:
            day_map[day] = col
    return day_map


def _detectar_layout_mensual(ws) -> tuple[int, int, int | None, Dict[int, int], int]:
    # 1) Detectar fila de encabezados por columna CONECTOR.
    header_row = None
    connector_col = None
    for row in range(1, min(ws.max_row, 40) + 1):
        candidate = _buscar_columna_por_aliases_ws(ws, row, ALIAS_COL_CONECTOR)
        if candidate is not None:
            header_row = row
            connector_col = candidate
            break

    if header_row is None or connector_col is None:
        raise ValueError(
            f"No se detecto columna CONECTOR en la hoja mensual '{ws.title}'."
        )

    comentario_col = _buscar_columna_por_aliases_ws(
        ws, header_row, ALIAS_COL_COMENTARIO_MENSUAL
    )

    # 2) Regla esperada: dias en fila 9 / datos desde fila 10.
    day_row = 9
    day_cols = _mapa_dias_en_fila(ws, day_row)

    # 3) Fallback robusto si el layout del archivo difiere.
    if len(day_cols) < 7:
        best_day_row = None
        best_day_cols: Dict[int, int] = {}
        for row in range(header_row, min(ws.max_row, header_row + 8) + 1):
            candidate = _mapa_dias_en_fila(ws, row)
            if len(candidate) > len(best_day_cols):
                best_day_cols = candidate
                best_day_row = row
        if best_day_row is None or len(best_day_cols) < 7:
            raise ValueError(
                f"No se detectaron columnas de dia (1..31) en hoja mensual '{ws.title}'."
            )
        day_row = best_day_row
        day_cols = best_day_cols

    data_start = max(day_row + 1, header_row + 1)
    return header_row, connector_col, comentario_col, day_cols, data_start


def _extraer_comentario_celda(cell) -> str:
    comment = getattr(cell, "comment", None)
    if comment is None:
        return ""
    return safe_str(getattr(comment, "text", ""))


def _tokenizar_estado(value: object) -> set[str]:
    normalized = normalize_text(value)
    if not normalized:
        return set()
    return {token for token in re.split(r"[^A-Z0-9]+", normalized) if token}


def _is_sig_rpmo_sheet(title: object) -> bool:
    return normalize_text(title) == "SIGRPMO"


def _seleccionar_hojas_mensual(wb, hoja_mensual: str | None = None):
    if hoja_mensual:
        for ws in wb.worksheets:
            if normalize_text(ws.title) == normalize_text(hoja_mensual):
                if _is_sig_rpmo_sheet(ws.title):
                    raise ValueError(
                        "La hoja mensual solicitada es SIGRPMO y no corresponde para etapa 3."
                    )
                _detectar_layout_mensual(ws)
                return [ws]
        raise ValueError(
            f"No se encontro hoja mensual '{hoja_mensual}'. "
            f"Hojas disponibles: {', '.join(wb.sheetnames)}"
        )

    selected = []
    for ws in wb.worksheets:
        if _is_sig_rpmo_sheet(ws.title):
            continue
        try:
            _detectar_layout_mensual(ws)
            selected.append(ws)
        except Exception:
            continue

    if not selected:
        raise ValueError(
            "No se encontro una hoja mensual valida (con CONECTOR y dias 1..31)."
        )
    return selected


def _es_mejor_estado(actual: Tuple[str, str], candidato: Tuple[str, str]) -> bool:
    estado_actual, comentario_actual = actual
    estado_candidato, comentario_candidato = candidato
    prio_actual = ESTADO_PRIORIDAD_ETAPA3.get(estado_actual, 0)
    prio_candidato = ESTADO_PRIORIDAD_ETAPA3.get(estado_candidato, 0)
    if prio_candidato > prio_actual:
        return True
    if prio_candidato < prio_actual:
        return False
    return len(comentario_candidato) > len(comentario_actual)


def _dias_para_hoja_en_rango(
    fechas_rango: List[date],
    sheet_year: int | None,
    sheet_month: int | None,
) -> List[int]:
    if not fechas_rango:
        return []
    if sheet_year is None or sheet_month is None:
        dias = sorted({f.day for f in fechas_rango})
        return [d for d in dias if 1 <= d <= 31]
    dias = sorted(
        {
            f.day
            for f in fechas_rango
            if f.year == sheet_year and f.month == sheet_month and 1 <= f.day <= 31
        }
    )
    return dias


def obtener_estado_y_comentario(
    ws,
    row: int,
    day_cols: Dict[int, int],
    dia_inicio: int,
    dia_fin: int,
    comentario_col: int | None = None,
    dias_permitidos: set[int] | None = None,
) -> Tuple[str, str]:
    """
    Evalua estado en rango dinamico (incluyente):
    for dia in range(dia_inicio, dia_fin + 1)
    Prioridad: S > NR > RR > R.
    """
    comentario_base = (
        safe_str(ws.cell(row=row, column=comentario_col).value)
        if comentario_col is not None
        else ""
    )

    mejor_estado = ""
    mejor_comentario = ""
    for dia in range(dia_inicio, dia_fin + 1):
        if dias_permitidos is not None and dia not in dias_permitidos:
            continue
        col = day_cols.get(dia)
        if col is None:
            continue

        cell = ws.cell(row=row, column=col)
        comentario_celda = _extraer_comentario_celda(cell)

        tokens = _tokenizar_estado(cell.value)
        valor_norm = normalize_text(cell.value)
        es_nr = "NR" in tokens or valor_norm in {"NO REALIZADO", "NOREALIZADO"}
        es_rr = "RR" in tokens
        es_r = ("R" in tokens or valor_norm == "R") and not es_nr and not es_rr

        # Regla solicitada: si mensual marca R, clasificar como R/TT
        # aunque exista comentario en celda.
        if es_r:
            candidato = ("R", comentario_celda or comentario_base)
        elif comentario_celda:
            candidato = ("S", comentario_celda)
        elif es_nr:
            candidato = ("NR", comentario_base)
        elif es_rr:
            candidato = ("RR", comentario_base)
        else:
            continue

        if _es_mejor_estado((mejor_estado, mejor_comentario), candidato):
            mejor_estado, mejor_comentario = candidato

    return mejor_estado, mejor_comentario


def leer_mensual(
    path_programa_mensual: Path | str,
    hoja_mensual: str | None = None,
) -> Dict[str, Any]:
    wb = load_workbook(Path(path_programa_mensual), data_only=False, read_only=False)
    hojas = _seleccionar_hojas_mensual(wb, hoja_mensual=hoja_mensual)

    sheets_data: List[Dict[str, Any]] = []
    total_connectors = 0
    for ws in hojas:
        _, connector_col, comentario_col, day_cols, data_start = _detectar_layout_mensual(ws)
        period = _periodo_desde_titulo_hoja_mensual(ws.title)
        sheet_year = period[0] if period else None
        sheet_month = period[1] if period else None

        rows_by_connector: Dict[str, List[int]] = {}
        blank_streak = 0
        for row in range(data_start, ws.max_row + 1):
            conector = safe_str(ws.cell(row=row, column=connector_col).value)
            if not conector:
                blank_streak += 1
                if blank_streak >= 300:
                    break
                continue
            blank_streak = 0
            key_conector = _normalizar_conector(conector)
            if not key_conector:
                continue
            rows_by_connector.setdefault(key_conector, []).append(row)

        total_connectors += len(rows_by_connector)
        sheets_data.append(
            {
                "ws": ws,
                "sheet_name": ws.title,
                "year": sheet_year,
                "month": sheet_month,
                "comentario_col": comentario_col,
                "day_cols": day_cols,
                "rows_by_connector": rows_by_connector,
            }
        )

    print(
        "[Etapa 3] Mensual leido: "
        f"{len(sheets_data)} hojas validas, "
        f"{total_connectors} conectores unicos (agregado)."
    )
    return {
        "workbook": wb,
        "sheets": sheets_data,
    }


def _leer_diccionario_clasificacion(
    path_diccionario: Path | str,
) -> List[Tuple[str, str]]:
    wb = load_workbook(Path(path_diccionario), data_only=True, read_only=True)
    try:
        keyword_col = None
        codigo_col = None
        header_row = None
        target_ws = None

        for ws in wb.worksheets:
            for row in range(1, min(ws.max_row, 40) + 1):
                k_col = _buscar_columna_por_aliases_ws(ws, row, ALIAS_COL_DICCIONARIO_PALABRA)
                c_col = _buscar_columna_por_aliases_ws(ws, row, ALIAS_COL_DICCIONARIO_CODIGO)
                if k_col is not None and c_col is not None:
                    keyword_col = k_col
                    codigo_col = c_col
                    header_row = row
                    target_ws = ws
                    break
            if target_ws is not None:
                break

        if target_ws is None or header_row is None or keyword_col is None or codigo_col is None:
            raise ValueError(
                "No se encontraron columnas 'PALABRA CLAVE COMENTARIO' y "
                "'CODIGO DE CIERRE' en el diccionario."
            )

        rules: List[Tuple[str, str]] = []
        blank_streak = 0
        for row in range(header_row + 1, target_ws.max_row + 1):
            keyword_raw = safe_str(target_ws.cell(row=row, column=keyword_col).value)
            codigo = safe_str(target_ws.cell(row=row, column=codigo_col).value).upper()
            if not keyword_raw and not codigo:
                blank_streak += 1
                if blank_streak >= 200:
                    break
                continue
            blank_streak = 0
            if not codigo:
                continue
            rules.append((normalize_text(keyword_raw), codigo))

    finally:
        wb.close()

    # Priorizamos keyword mas larga (mas especifica); keyword vacia queda al final.
    rules.sort(key=lambda item: (len(item[0]), item[0]), reverse=True)
    print(f"[Etapa 3] Diccionario leido: {len(rules)} reglas.")
    return rules


def clasificar_diccionario(
    comentario: object,
    diccionario: List[Tuple[str, str]],
) -> str:
    comentario_norm = normalize_text(comentario)
    fallback = ""
    for keyword_norm, codigo in diccionario:
        code = safe_str(codigo).upper()
        if not code:
            continue
        if not keyword_norm:
            if not fallback:
                fallback = code
            continue
        if keyword_norm in comentario_norm:
            return code
    return fallback


def clasificar_fila(
    grupo_trab: object,
    estado: object,
    comentario: object,
    diccionario: List[Tuple[str, str]],
) -> str:
    if normalize_text(grupo_trab) not in GRUPOS_TRABAJO_ETAPA3:
        return ""

    estado_norm = normalize_text(estado)
    if estado_norm == "R":
        return "TT"
    if estado_norm in {"NR", "RR"}:
        return "MI"
    if estado_norm == "S":
        return clasificar_diccionario(comentario=comentario, diccionario=diccionario)
    return ""


def actualizar_cierre_ot(
    path_cierre_ot_base: Path | str,
    mensual_lookup: Dict[str, Any],
    diccionario: List[Tuple[str, str]],
    output_path: Path | str,
) -> Dict[str, Any]:
    mensual_sheets: List[Dict[str, Any]] = mensual_lookup.get("sheets", [])

    wb = load_workbook(Path(path_cierre_ot_base))
    stats = {
        "hojas_totales": len(wb.worksheets),
        "hojas_procesadas": 0,
        "hojas_omitidas_sig_rpmo": 0,
        "filas_evaluadas": 0,
        "filas_con_match_mensual": 0,
        "codigos_actualizados": 0,
        "comentarios_actualizados": 0,
        "hojas_mensual_usadas": len(mensual_sheets),
    }
    try:
        for ws in wb.worksheets:
            if _is_sig_rpmo_sheet(ws.title):
                stats["hojas_omitidas_sig_rpmo"] += 1
                continue

            stats["hojas_procesadas"] += 1
            day = _dia_desde_nombre_hoja(ws.title)
            if day is None:
                continue

            col_desc2 = _buscar_columna_por_aliases_ws(ws, 1, ALIAS_COL_CIERRE_DESC2)
            col_matriz = _buscar_columna_por_aliases_ws(ws, 1, ALIAS_COL_CIERRE_MATRIZ)
            col_grupo = _buscar_columna_por_aliases_ws(ws, 1, ALIAS_COL_CIERRE_GRUPO)
            col_codigo = _buscar_columna_por_aliases_ws(ws, 1, ALIAS_COL_CIERRE_CODIGO)
            col_comentario_adiciona = _buscar_columna_por_aliases_ws(
                ws, 1, ALIAS_COL_CIERRE_COMENTARIO_ADICIONA
            )
            col_fecha_inicio = _buscar_columna_por_aliases_ws(
                ws, 1, ALIAS_COL_CIERRE_FECHA_INICIO
            )
            col_fecha_fin = _buscar_columna_por_aliases_ws(ws, 1, ALIAS_COL_CIERRE_FECHA_FIN)

            if col_desc2 is None or col_matriz is None or col_grupo is None:
                continue
            if col_codigo is None and col_comentario_adiciona is None:
                continue

            blank_streak = 0
            for row in range(2, ws.max_row + 1):
                desc2 = safe_str(ws.cell(row=row, column=col_desc2).value)
                matriz = safe_str(ws.cell(row=row, column=col_matriz).value)
                if not desc2 and not matriz:
                    blank_streak += 1
                    if blank_streak >= 250:
                        break
                    continue
                blank_streak = 0

                stats["filas_evaluadas"] += 1
                grupo = safe_str(ws.cell(row=row, column=col_grupo).value)
                if normalize_text(grupo) not in GRUPOS_TRABAJO_ETAPA3:
                    continue

                llave = _normalizar_llave(desc2, matriz)
                if not llave:
                    continue

                fecha_inicio = (
                    _to_date(ws.cell(row=row, column=col_fecha_inicio).value)
                    if col_fecha_inicio is not None
                    else None
                )
                fecha_fin = (
                    _to_date(ws.cell(row=row, column=col_fecha_fin).value)
                    if col_fecha_fin is not None
                    else None
                )
                fechas_rango = _iter_date_range(fecha_inicio, fecha_fin)

                estado_mejor = ""
                comentario_mejor = ""
                for sheet_data in mensual_sheets:
                    rows = sheet_data.get("rows_by_connector", {}).get(llave, [])
                    if not rows:
                        continue
                    dias_validos = _dias_para_hoja_en_rango(
                        fechas_rango,
                        sheet_data.get("year"),
                        sheet_data.get("month"),
                    )
                    if not dias_validos:
                        if fechas_rango:
                            continue
                        if day is None:
                            continue
                        dias_validos = [day]

                    dia_inicio = min(dias_validos)
                    dia_fin = max(dias_validos)
                    for mensual_row in rows:
                        estado_row, comentario_row = obtener_estado_y_comentario(
                            sheet_data["ws"],
                            mensual_row,
                            sheet_data["day_cols"],
                            dia_inicio,
                            dia_fin,
                            comentario_col=sheet_data.get("comentario_col"),
                            dias_permitidos=set(dias_validos),
                        )
                        if not estado_row:
                            continue
                        if _es_mejor_estado(
                            (estado_mejor, comentario_mejor),
                            (estado_row, comentario_row),
                        ):
                            estado_mejor = estado_row
                            comentario_mejor = comentario_row

                if not estado_mejor:
                    continue

                stats["filas_con_match_mensual"] += 1
                codigo = clasificar_fila(
                    grupo_trab=grupo,
                    estado=estado_mejor,
                    comentario=comentario_mejor,
                    diccionario=diccionario,
                )

                if col_codigo is not None and codigo:
                    ws.cell(row=row, column=col_codigo).value = codigo
                    stats["codigos_actualizados"] += 1

                if col_comentario_adiciona is not None and comentario_mejor:
                    ws.cell(row=row, column=col_comentario_adiciona).value = comentario_mejor
                    stats["comentarios_actualizados"] += 1

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
    finally:
        wb.close()

    stats["output_path"] = str(Path(output_path))
    print(
        "[Etapa 3] Cierre OT actualizado: "
        f"codigos={stats['codigos_actualizados']} | "
        f"comentarios={stats['comentarios_actualizados']} | "
        f"hojas={stats['hojas_procesadas']} (omitida SIGRPMO={stats['hojas_omitidas_sig_rpmo']})"
    )
    return stats


def etapa3_clasificacion(
    path_cierre_ot_base: Path | str,
    path_programa_mensual: Path | str,
    path_diccionario: Path | str,
    output_path: Path | str,
    hoja_mensual: str | None = None,
) -> Dict[str, Any]:
    diccionario = _leer_diccionario_clasificacion(path_diccionario)
    mensual = leer_mensual(path_programa_mensual, hoja_mensual=hoja_mensual)
    try:
        return actualizar_cierre_ot(
            path_cierre_ot_base=path_cierre_ot_base,
            mensual_lookup=mensual,
            diccionario=diccionario,
            output_path=output_path,
        )
    finally:
        wb = mensual.get("workbook")
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
