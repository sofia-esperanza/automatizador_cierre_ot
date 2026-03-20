from __future__ import annotations

import datetime as dt
import unicodedata
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

TURN_START_WEEKDAY = 3  # jueves

NORMAL_HEADERS = [
    "Referencia",
    "Descripción",
    "Descripción de tarea programada 2",
    "Fecha de inicio del plan",
    "Fecha de finalización planificada",
    "Grupo trab",
    "Observación",
    "CODIGO DE CIERRE",
    "Descripción, obs",
    "Descripción Descripción de tarea programada 2",
    "Comentario adiciona",
    "Matriz Terreno",
]

SIG_HEADERS = [
    "Referencia",
    "Descripción",
    "Descripción de tarea programada 2",
    "Fecha de inicio del plan",
    "Fecha de finalización planificada",
    "Grupo trab",
    "CODIGO DE CIERRE",
    "Descripción, obs",
    "Descripción Descripción de tarea programada 2",
    "Comentario adiciona",
]

DESC_MAP = {
    "MEDICION DE NIVELES FREATICOS (PAT EIA)": "MED DE NIVELES FREATICOS",
    "MUESTREO MANUAL DE AGUAS SUBTERRANEAS": "MM DE AGUAS SUBTERRÁNEAS",
    "MUESTREO MANUAL DE AGUAS SUPERFICIALES": "MM DE AGUAS SUPERFICIALES",
    "OPERACION DE ESTACIONES CONTINUAS": "OP DE ESTACIONES CONTINUAS",
    "MANTENCION DE ESTACIONES CONTINUAS": "MANT DE ESTACIONES CONTINUAS",
    "MEDICION CON OLFATOMETRO DE CAMPO": "MED CON OLFATÓMETRO DE CAMPO",
    "OPERACION ESTACIONES METEOROLOGICAS": "Op Estaciones Meteorologicas",
    "MONITOREO EN SALAS DE ESTACION Y TK": "MONIT EN SALAS DE ESTACION Y TK",
    "TRASLADO PERS. Y CAM. TURNO CONTRATOS": "TRAS PERS. Y CAM. TURNO CONTRAT",
    "TRASLADO MUESTRAS Y MATERIAL MUESTREO": "TRAS MUESTRAS Y MATERIAL MUEST",
    "MUESTREO M. AGUAS SUPERF. (FOTOMETRO)": "MM. AGUAS  SUPERF. (FOTOMETRO)",
}

MATRIZ_TERRENO_MAP = {
    "MED DE NIVELES FREATICOS": "NF",
    "MEDICION DE NIVELES FREATICOS": "NF",
    "MEDICION NIVELES FREATICOS": "NF",
    "MM DE AGUAS SUBTERRANEAS": "ASUB",
    "MM DE AGUAS SUPERFICIALES": "ASUP",
    "MONITOREO DE BANOS Y CASINOS": "AP",
    "MONIT EN SALAS DE ESTACION Y TK": "AP",
    "MM. AGUAS  SUPERF. (FOTOMETRO)": "FOTOMETRO",
    "MUESTREO M. AGUAS SUPERF (FOTOMETRO)": "FOTOMETRO",
    "MEDICION DE CAUDALES": "CAUDAL",
    "MEDICION CAUDALES": "CAUDAL",
    "MUESTREO DE RIL-AS (CALIDAD)": "AR",
    "MUESTREO PUNTUAL DE RIL-AS": "AR",
    "MUESTREO PUNTUAL DE RIL AS": "AR",
    "HOUSEKEEPING": "HK",
    "MONITOREO AGUAS SERVIDAS": "AR",
}

MSEWJO_COLUMN_MAP = {
    "Referencia": "D",
    "Descripción": "E",
    "Descripción de tarea programada 2": "HS",
    "Fecha de inicio del plan": "BD",
    "Fecha de finalización planificada": "BF",
    "Grupo trab": "CJ",
}


def parse_date(value: Any) -> Optional[dt.date]:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        try:
            return dt.date(1899, 12, 30) + dt.timedelta(days=int(value))
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y", "%d.%m.%y"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        try:
            return dt.date.fromisoformat(s)
        except Exception:
            return None
    return None


def get_turn_window(anchor_date: dt.date) -> tuple[dt.date, dt.date]:
    delta = (anchor_date.weekday() - TURN_START_WEEKDAY) % 7
    start = anchor_date - dt.timedelta(days=delta)
    end = start + dt.timedelta(days=6)
    return start, end


def _norm_key(text: str) -> str:
    compact = " ".join((text or "").strip().split())
    normalized = unicodedata.normalize("NFD", compact)
    no_accents = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return no_accents.upper()


def normalize_desc_b(text: Any) -> Any:
    if text is None:
        return None
    s = " ".join(str(text).strip().split())
    normalized = _norm_key(s)
    for source, target in DESC_MAP.items():
        if _norm_key(source) == normalized:
            return target
    return s


def clasificar_matriz_terreno(descripcion: Any, grupo: Any) -> str:
    if not descripcion:
        return ""
    grupo_norm = _norm_key(str(grupo or ""))
    if grupo_norm not in {"SIGVA", "SIGVANC"}:
        return ""
    descripcion_norm = _norm_key(str(descripcion))
    for source, code in MATRIZ_TERRENO_MAP.items():
        if _norm_key(source) == descripcion_norm:
            return code
    return ""


def safe_sheet_name(date_value: dt.date, existing: set[str]) -> str:
    name = str(date_value.day)
    if name in existing:
        name = date_value.strftime("%d-%m")
    i = 2
    base = name
    while name in existing:
        name = f"{base}_{i}"
        i += 1
    return name[:31]


def apply_comment_rules(ws, comment_col_letter: str, last_row: int) -> None:
    dv = DataValidation(
        type="textLength", operator="lessThanOrEqual", formula1="45", allow_blank=False
    )
    dv.add(f"{comment_col_letter}1:{comment_col_letter}{last_row}")
    ws.add_data_validation(dv)

    dxf = DifferentialStyle(
        font=Font(b=True, i=True, strike=True),
        fill=PatternFill(patternType="solid", fgColor="FFFF0000"),
    )
    rule = Rule(
        type="expression",
        dxf=dxf,
        formula=[f'AND({comment_col_letter}2<>"", LEN({comment_col_letter}2)>50)'],
    )
    ws.conditional_formatting.add(f"{comment_col_letter}2:{comment_col_letter}{last_row}", rule)


def apply_table_borders_and_colors(ws, last_row: int, last_col: int, color_cols=("H", "I", "J")) -> None:
    thin = Side(style="thin", color="FF666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_map = {
        "H": PatternFill(patternType="solid", fgColor="FF9CE2DD"),
        "I": PatternFill(patternType="solid", fgColor="FF00B0AA"),
        "J": PatternFill(patternType="solid", fgColor="FFB8B8B8"),
    }

    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            ws.cell(row, col).border = border

    for col_letter in color_cols:
        col_idx = column_index_from_string(col_letter)
        fill = fill_map.get(col_letter)
        if fill is None:
            continue
        for row in range(2, last_row + 1):
            ws.cell(row, col_idx).fill = fill

    for col_letter in ("G", "H", "I", "J"):
        col_idx = column_index_from_string(col_letter)
        for row in range(2, last_row + 1):
            cell = ws.cell(row, col_idx)
            font = copy(cell.font)
            font.bold = True
            cell.font = font


def extract_rows_from_msewjo(path_msewjo: str | Path) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(Path(path_msewjo), data_only=False, read_only=True)
    try:
        ws = wb.active
        col_idx = {k: column_index_from_string(v) for k, v in MSEWJO_COLUMN_MAP.items()}
        min_col = min(col_idx.values())
        max_col = max(col_idx.values())

        def value_at(row_cells, absolute_col: int):
            i = absolute_col - min_col
            if i < 0 or i >= len(row_cells):
                return None
            return row_cells[i].value

        rows: List[Dict[str, Any]] = []
        for row_cells in ws.iter_rows(min_row=2, min_col=min_col, max_col=max_col):
            ref = value_at(row_cells, col_idx["Referencia"])
            if ref is None or str(ref).strip() == "":
                continue

            item = {
                "Referencia": ref,
                "Descripción": value_at(row_cells, col_idx["Descripción"]),
                "Descripción de tarea programada 2": value_at(
                    row_cells, col_idx["Descripción de tarea programada 2"]
                ),
                "Fecha de inicio del plan": value_at(row_cells, col_idx["Fecha de inicio del plan"]),
                "Fecha de finalización planificada": value_at(
                    row_cells, col_idx["Fecha de finalización planificada"]
                ),
                "Grupo trab": value_at(row_cells, col_idx["Grupo trab"]),
            }
            item["_bf_date"] = parse_date(item["Fecha de finalización planificada"])
            rows.append(item)
        return rows
    finally:
        wb.close()


def build_output_workbook(
    rows: List[Dict[str, Any]],
    turn_start: Optional[dt.date] = None,
    turn_end: Optional[dt.date] = None,
    include_sin_fecha: bool = True,
) -> openpyxl.Workbook:
    if (turn_start is None) != (turn_end is None):
        raise ValueError("turn_start y turn_end deben definirse juntos.")
    if turn_start is not None and turn_end is not None and turn_start > turn_end:
        raise ValueError("turn_start no puede ser mayor que turn_end.")

    def in_turn(date_value: Optional[dt.date]) -> bool:
        if date_value is None or turn_start is None or turn_end is None:
            return False
        return turn_start <= date_value <= turn_end

    sig_all = [r for r in rows if _norm_key(str(r.get("Grupo trab") or "")) == "SIGRPMO"]
    normal_all = [r for r in rows if _norm_key(str(r.get("Grupo trab") or "")) != "SIGRPMO"]

    if turn_start is not None and turn_end is not None:
        sig = [r for r in sig_all if in_turn(r.get("_bf_date"))]
        normal = [r for r in normal_all if in_turn(r.get("_bf_date"))]
        sin_fecha = [r for r in normal_all if r.get("_bf_date") is None]
    else:
        sig = sig_all
        normal = normal_all
        sin_fecha = [r for r in normal_all if r.get("_bf_date") is None]

    groups = defaultdict(list)
    for row in normal:
        bf_date = row.get("_bf_date")
        if bf_date is not None:
            groups[bf_date].append(row)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFFFF")
    header_font_black = Font(bold=True, color="FF000000")
    header_fill_default = PatternFill(patternType="solid", fgColor="FF297A76")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(vertical="top", wrap_text=True)

    header_fill_special = {
        8: PatternFill(patternType="solid", fgColor="FF00B0AA"),
        9: PatternFill(patternType="solid", fgColor="FF297A76"),
        10: PatternFill(patternType="solid", fgColor="FF666666"),
    }

    existing = set()

    def setup_headers(ws, headers) -> None:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill_special.get(col, header_fill_default)
            cell.font = header_font_black if col == 8 else header_font
            cell.alignment = header_align

    for date_value in sorted(groups.keys()):
        sheet_name = safe_sheet_name(date_value, existing)
        existing.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        setup_headers(ws, NORMAL_HEADERS)

        data_rows = groups[date_value]
        for i, item in enumerate(data_rows, start=2):
            ws.cell(i, 1, item["Referencia"]).alignment = data_align
            ws.cell(i, 2, normalize_desc_b(item["Descripción"])).alignment = data_align
            ws.cell(i, 3, item["Descripción de tarea programada 2"]).alignment = data_align
            ws.cell(i, 4, item["Fecha de inicio del plan"]).alignment = data_align
            ws.cell(i, 5, item["Fecha de finalización planificada"]).alignment = data_align
            ws.cell(i, 6, item["Grupo trab"]).alignment = data_align
            ws.cell(i, 8, "").alignment = data_align
            ws.cell(i, 7).value = (
                f'=IF(H{i}="TT","TRABAJO TERMINADO",IF(H{i}="MI","TRABAJO NO TERMINADO",""))'
            )
            ws.cell(i, 9).value = (
                f'=IF(B{i}="", G{i}, IF(G{i}="", LEFT(B{i},45), '
                f'IF(45-LEN(G{i})-2<=0, G{i}, LEFT(B{i},45-LEN(G{i})-2)&", "&G{i})))'
            )
            ws.cell(i, 10).value = (
                f'=IF(H{i}="MI","NO REALIZADO " & B{i} & " EN " & C{i} & '
                f'IF(K{i}<>""," " & K{i},""),IF(H{i}="TT","SE REALIZA " & B{i} & " EN " & C{i} & '
                f'IF(K{i}<>""," " & K{i},""),""))'
            )
            ws.cell(i, 11, None).alignment = data_align
            ws.cell(i, 12, clasificar_matriz_terreno(normalize_desc_b(item["Descripción"]), item["Grupo trab"])).alignment = data_align

        last_row = max(2, 1 + len(data_rows))
        apply_comment_rules(ws, "I", last_row)
        for col, width in {
            "A": 12,
            "B": 45,
            "C": 30,
            "D": 16,
            "E": 20,
            "F": 14,
            "G": 22,
            "H": 14,
            "I": 35,
            "J": 80,
            "K": 28,
        }.items():
            ws.column_dimensions[col].width = width
        apply_table_borders_and_colors(
            ws, last_row=last_row, last_col=len(NORMAL_HEADERS), color_cols=("H", "I", "J")
        )

    if include_sin_fecha and sin_fecha:
        ws = wb.create_sheet(title="SIN_FECHA")
        setup_headers(ws, NORMAL_HEADERS)
        for i, item in enumerate(sin_fecha, start=2):
            ws.cell(i, 1, item["Referencia"])
            ws.cell(i, 2, item["Descripción"])
            ws.cell(i, 3, item["Descripción de tarea programada 2"])
            ws.cell(i, 4, item["Fecha de inicio del plan"])
            ws.cell(i, 5, item["Fecha de finalización planificada"])
            ws.cell(i, 6, item["Grupo trab"])
            ws.cell(i, 8, "")
            ws.cell(i, 7).value = (
                f'=IF(H{i}="TT","TRABAJO TERMINADO",IF(H{i}="MI","TRABAJO NO TERMINADO",""))'
            )
            ws.cell(i, 9).value = (
                f'=IF(B{i}="", G{i}, IF(G{i}="", LEFT(B{i},45), '
                f'IF(45-LEN(G{i})-2<=0, G{i}, LEFT(B{i},45-LEN(G{i})-2)&", "&G{i})))'
            )
            ws.cell(i, 10).value = (
                f'=IF(H{i}="MI","NO REALIZADO " & B{i} & " EN " & C{i} & '
                f'IF(K{i}<>""," " & K{i},""),IF(H{i}="TT","SE REALIZA " & B{i} & " EN " & C{i} & '
                f'IF(K{i}<>""," " & K{i},""),""))'
            )

        last_row = max(2, 1 + len(sin_fecha))
        apply_comment_rules(ws, "I", last_row)
        for col, width in {
            "A": 12,
            "B": 45,
            "C": 30,
            "D": 16,
            "E": 20,
            "F": 14,
            "G": 22,
            "H": 14,
            "I": 35,
            "J": 80,
            "K": 28,
        }.items():
            ws.column_dimensions[col].width = width
        apply_table_borders_and_colors(
            ws, last_row=last_row, last_col=len(NORMAL_HEADERS), color_cols=("H", "I", "J")
        )

    if sig:
        ws = wb.create_sheet(title="SIGRPMO")
        setup_headers(ws, SIG_HEADERS)

        for i, item in enumerate(sig, start=2):
            ws.cell(i, 1, item["Referencia"])
            ws.cell(i, 2, item["Descripción"])
            ws.cell(i, 3, item["Descripción de tarea programada 2"])
            ws.cell(i, 4, item["Fecha de inicio del plan"])
            ws.cell(i, 5, item["Fecha de finalización planificada"])
            ws.cell(i, 6, item["Grupo trab"])
            ws.cell(i, 7, "")
            ws.cell(i, 8).value = f'=IF(G{i}="TT", B{i} & " EMITIDO", IF(G{i}="MI", B{i} & " NO EMITIDO", ""))'
            ws.cell(i, 9).value = (
                f'=IF(G{i}="TT", B{i} &" EMITIDO dentro de plazo, según planificación ", '
                f'IF(G{i}="MI", B{i} & "NO SE RECIBE REPORTE EN LA FECHA SOLICITADA", ""))'
            )

        last_row = max(2, 1 + len(sig))
        apply_comment_rules(ws, "I", last_row)
        for col, width in {
            "A": 12,
            "B": 45,
            "C": 40,
            "D": 16,
            "E": 20,
            "F": 14,
            "G": 14,
            "H": 35,
            "I": 60,
            "J": 28,
        }.items():
            ws.column_dimensions[col].width = width
        apply_table_borders_and_colors(
            ws, last_row=last_row, last_col=len(SIG_HEADERS), color_cols=("G", "H", "I")
        )

    def sheet_sort_key(name: str):
        if name == "SIGRPMO":
            return (2, 999)
        if name.startswith("SIN_FECHA"):
            return (1, 999)
        if name.isdigit():
            return (0, int(name))
        return (0, 998)

    wb._sheets.sort(key=lambda s: sheet_sort_key(s.title))
    return wb


def generar_cierre_turno_desde_msewjo(
    path_msewjo: str | Path,
    output_path: str | Path,
    turn_start: Optional[dt.date] = None,
    turn_end: Optional[dt.date] = None,
    include_sin_fecha: bool = True,
) -> Dict[str, Any]:
    rows = extract_rows_from_msewjo(path_msewjo)
    wb = build_output_workbook(
        rows,
        turn_start=turn_start,
        turn_end=turn_end,
        include_sin_fecha=include_sin_fecha,
    )
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()
    return {"output_path": output_file, "rows_count": len(rows)}
