from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from procesos.actualizar_programa_mensual import _find_header_columns, _canonical_matriz
from utils.texto_utils import safe_str, normalize_key

print("Iniciando diagnóstico...")

base = Path.cwd() / "output" / "_temp" / "etapa_2_actualizacion_mensual"
no = pd.read_excel(base / "no_cruzados.xlsx")

mensual = Path.home() / "Downloads" / "PROG & OT TURNOS" / "TURNOSS" / "19-25 MARZO" / "PROGRAMA DE MONITOREO MARZO 26.xlsx"
wb = load_workbook(mensual, data_only=True, read_only=True)
ws = wb["MARZO 2026"]

codigo_col, matriz_col, day_to_col, data_start_row, max_row = _find_header_columns(ws)

matrices_por_codigo = {}
row_lookup = {}
blank = 0
for r in range(data_start_row, max_row + 1):
    codigo = safe_str(ws.cell(r, codigo_col).value)
    matriz = safe_str(ws.cell(r, matriz_col).value)
    if not codigo and not matriz:
        blank += 1
        if blank >= 200 and r > data_start_row + 200:
            break
        continue
    blank = 0
    if codigo:
        matrices_por_codigo.setdefault(codigo, set()).add(matriz)
    key = normalize_key(codigo, _canonical_matriz(matriz))
    row_lookup[key] = r

diag = []
for _, rec in no.iterrows():
    codigo = safe_str(rec["CODIGO"])
    matriz = safe_str(rec["MATRIZ"])
    dia = int(rec["DIA"]) if pd.notna(rec["DIA"]) else None
    key = normalize_key(codigo, _canonical_matriz(matriz))

    diag.append({
        "CODIGO": codigo,
        "MATRIZ_SEMANAL": matriz,
        "DIA": dia,
        "ESTADO": safe_str(rec["ESTADO"]),
        "MOTIVO_ORIGINAL": safe_str(rec["MOTIVO"]),
        "CODIGO_EXISTE_EN_MENSUAL": codigo in matrices_por_codigo,
        "MATRICES_EN_MENSUAL_PARA_CODIGO": " | ".join(sorted(matrices_por_codigo.get(codigo, []))),
        "COINCIDE_CODIGO_MATRIZ": key in row_lookup,
        "COLUMNA_DIA_EXISTE": dia in day_to_col if dia is not None else False,
    })

out = base / "diagnostico_no_cruzados.xlsx"
pd.DataFrame(diag).to_excel(out, index=False)
wb.close()
print(f"Generado: {out}")
