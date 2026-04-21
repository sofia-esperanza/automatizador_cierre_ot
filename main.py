from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Dict

import openpyxl
import pandas as pd

from procesos.actualizar_programa_mensual import actualizar_programa_mensual
from procesos.clasificar_cierre_ot import etapa3_clasificacion
from procesos.generar_cierre_ot import generar_cierre_ot_base
from procesos.generar_cierre_turno_desde_msewjo import (
    generar_cierre_turno_desde_msewjo,
    get_turn_window,
)
from procesos.limpieza_msewjo import limpiar_msewjo
from utils.excel_utils import save_dataframe_to_excel
from utils.texto_utils import normalize_column_name, normalize_text, safe_str

TEMP_DIRNAME = "_temp"
ETAPA_1_DIRNAME = "etapa_1_limpieza_base"
ETAPA_2_DIRNAME = "etapa_2_actualizacion_mensual"
ETAPA_3_DIRNAME = "etapa_3_clasificacion"
ETAPA_2_INTERNO_DIRNAME = "_interno"
ESTADOS_VALIDOS_MENSUAL = {"R", "NR", "RR", "S"}
GRUPOS_TRABAJO_CLASIFICABLES = {"SIGVA", "SIGVANC"}
MESES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def _is_sig_rpmo_sheet(title: object) -> bool:
    return normalize_text(title) == "SIGRPMO"


def _stage_dir(carpeta_salida: str | Path, stage_dirname: str) -> Path:
    stage_dir = Path(carpeta_salida) / TEMP_DIRNAME / stage_dirname
    stage_dir.mkdir(parents=True, exist_ok=True)
    return stage_dir


def _eliminar_archivos_etapa_3_no_deseados(stage_dir: Path) -> None:
    for nombre in [
        "cierre_ot_final.xlsx",
        "registros_clasificados.xlsx",
        "registros_turno_aplicado_etapa_3.xlsx",
    ]:
        ruta = stage_dir / nombre
        if ruta.exists():
            try:
                ruta.unlink()
            except Exception:
                pass


def _parse_fecha_usuario(value: str | dt.date | None, campo: str) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    text = str(value).strip()
    if not text:
        return None

    formatos = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]
    for formato in formatos:
        try:
            return dt.datetime.strptime(text, formato).date()
        except ValueError:
            continue

    raise ValueError(
        f"Fecha invalida en '{campo}': '{text}'. Usa formato dd/mm/aaaa."
    )


def _normalize_matriz_for_key(value: object) -> str:
    normalized = normalize_text(value)
    if normalized in {"AGUA POTABLE"}:
        return "AP"
    if normalized in {"AGUA RESIDUAL"}:
        return "AR"
    if normalized in {"AGUA SUBTERRANEA", "AGUAS SUBTERRANEAS"}:
        return "ASUB"
    if normalized in {"AGUA SUPERFICIAL", "AGUAS SUPERFICIALES"}:
        return "ASUP"
    if normalized in {"NIVEL FREATICO", "NIVELES FREATICOS"}:
        return "NF"
    return normalized


def _coerce_day(value: object) -> int | None:
    if isinstance(value, int) and 1 <= value <= 31:
        return value
    if isinstance(value, float) and value.is_integer():
        day = int(value)
        return day if 1 <= day <= 31 else None
    if isinstance(value, dt.datetime):
        return value.day
    if isinstance(value, dt.date):
        return value.day
    text = safe_str(value)
    if text.isdigit():
        day = int(text)
        return day if 1 <= day <= 31 else None
    return None


def _parse_date_cell(value: object) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def _sheet_day_from_title(title: str) -> int | None:
    token = safe_str(title).split("-")[0]
    if token.isdigit():
        day = int(token)
        return day if 1 <= day <= 31 else None
    return None


def _extract_punto_from_tarea(value: object) -> str:
    text = safe_str(value)
    if not text:
        return ""
    first = text.split()[0]
    if "/" in first:
        first = first.split("/")[0]
    return first.strip().strip(",;")


def _resolver_cierre_turno_base_path(
    carpeta_salida: str | Path,
    fecha_desde: dt.date | None = None,
    fecha_hasta: dt.date | None = None,
) -> Path:
    stage_dir = Path(carpeta_salida) / TEMP_DIRNAME / ETAPA_1_DIRNAME
    candidatos = sorted(stage_dir.glob("Cierre de OT Turno *.xlsx"))
    if not candidatos:
        raise FileNotFoundError(
            f"No se encontro 'Cierre de OT Turno *.xlsx' en {stage_dir}. Ejecuta etapa 1 primero."
        )

    if fecha_desde is not None and fecha_hasta is not None:
        periodo = f"{fecha_desde.strftime('%d.%m')} al {fecha_hasta.strftime('%d.%m')}"
        exactos = [p for p in candidatos if periodo in p.name]
        if exactos:
            return max(exactos, key=lambda p: p.stat().st_mtime)
    return max(candidatos, key=lambda p: p.stat().st_mtime)


def _leer_registros_turno_desde_etapa1(path_cierre_turno_base: Path | str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(Path(path_cierre_turno_base), data_only=True, read_only=True)
    rows: list[dict[str, object]] = []
    try:
        for ws in wb.worksheets:
            if normalize_text(ws.title) == "SIGRPMO":
                continue
            day_sheet = _sheet_day_from_title(ws.title)
            if day_sheet is None:
                continue

            header_map = {
                normalize_column_name(ws.cell(row=1, column=c).value): c
                for c in range(1, ws.max_column + 1)
            }
            col_desc = header_map.get(normalize_column_name("Descripción"))
            col_tarea = header_map.get(normalize_column_name("Descripción de tarea programada 2"))
            col_fecha = header_map.get(normalize_column_name("Fecha de finalización planificada"))
            col_matriz = header_map.get(normalize_column_name("Matriz Terreno"))
            # Refuerza detección de encabezados cuando el archivo trae variantes sin acento.
            col_desc = col_desc or header_map.get(normalize_column_name("Descripcion"))
            col_tarea = col_tarea or header_map.get(normalize_column_name("Descripcion de tarea programada 2"))
            col_fecha = col_fecha or header_map.get(normalize_column_name("Fecha de finalizacion planificada"))
            col_grupo_trab = header_map.get(normalize_column_name("Grupo trab"))
            if col_tarea is None or col_matriz is None:
                continue

            for r in range(2, ws.max_row + 1):
                tarea = ws.cell(row=r, column=col_tarea).value
                matriz = ws.cell(row=r, column=col_matriz).value
                punto = _extract_punto_from_tarea(tarea)
                tipo = safe_str(matriz)
                if not punto or not tipo:
                    continue
                fecha_ref = (
                    _parse_date_cell(ws.cell(row=r, column=col_fecha).value) if col_fecha else None
                )
                dia = day_sheet or (fecha_ref.day if fecha_ref else None)
                if dia is None:
                    continue
                descripcion = safe_str(ws.cell(row=r, column=col_desc).value) if col_desc else ""
                grupo_trab = (
                    safe_str(ws.cell(row=r, column=col_grupo_trab).value)
                    if col_grupo_trab is not None
                    else ""
                )
                rows.append(
                    {
                        "PUNTO": punto,
                        "TIPO": tipo,
                        "DIA": int(dia),
                        "COMENTARIO": descripcion,
                        "ACTIVIDAD": descripcion,
                        "DESC_TAREA_2": safe_str(tarea),
                        "GRUPO_TRAB": grupo_trab,
                        "FECHA_REFERENCIA": fecha_ref,
                        "ORIGEN_HOJA": ws.title,
                        "ORIGEN_FILA": r,
                    }
                )
    finally:
        wb.close()

    if not rows:
        raise ValueError(
            "No se pudieron extraer registros validos desde Cierre OT Turno base (etapa 1)."
        )
    return pd.DataFrame(rows)


def _filtrar_df_por_rango(df: pd.DataFrame, fecha_desde: dt.date, fecha_hasta: dt.date) -> pd.DataFrame:
    fecha_series = pd.to_datetime(df["FECHA_REFERENCIA"], errors="coerce")
    if fecha_series.notna().any():
        mask = fecha_series.ge(pd.Timestamp(fecha_desde)) & fecha_series.le(pd.Timestamp(fecha_hasta))
        return df[mask.fillna(False)].copy()
    mask_dia = pd.to_numeric(df["DIA"], errors="coerce").between(
        fecha_desde.day, fecha_hasta.day, inclusive="both"
    )
    return df[mask_dia.fillna(False)].copy()


def _select_monthly_sheet(wb, fecha_desde: dt.date | None, fecha_hasta: dt.date | None):
    expected = None
    if fecha_desde is not None and fecha_hasta is not None:
        if fecha_desde.year == fecha_hasta.year and fecha_desde.month == fecha_hasta.month:
            expected = f"{MESES_ES[fecha_desde.month]} {fecha_desde.year}"

    if expected:
        for ws in wb.worksheets:
            if _is_sig_rpmo_sheet(ws.title):
                continue
            if normalize_text(ws.title) == normalize_text(expected):
                return ws

    for ws in wb.worksheets:
        if _is_sig_rpmo_sheet(ws.title):
            continue
        headers = [normalize_column_name(ws.cell(row=7, column=c).value) for c in range(1, min(ws.max_column, 90) + 1)]
        if any(h == "CODIGO" for h in headers) and any(h == "MATRIZ_TERRENO" for h in headers):
            return ws

    for ws in wb.worksheets:
        if not _is_sig_rpmo_sheet(ws.title):
            return ws
    return wb.worksheets[0]


def _normalize_connector(value: object) -> str:
    text = normalize_text(value)
    text = text.replace(" ", "")
    text = text.replace("\\", "/")
    return text


def _is_grupo_clasificable(value: object) -> bool:
    return normalize_text(value) in GRUPOS_TRABAJO_CLASIFICABLES


def _find_monthly_columns(ws):
    best = None
    for row in range(1, min(40, ws.max_row) + 1):
        code_col = None
        matriz_col = None
        connector_col = None
        comentario_col = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            header = normalize_column_name(val)
            if code_col is None and header in {"CODIGO", "CODIGO_MST", "CODIGO_PUNTO"}:
                code_col = col
            if matriz_col is None and header == "MATRIZ_TERRENO":
                matriz_col = col
            if connector_col is None and header == "CONECTOR":
                connector_col = col
            if comentario_col is None and header in {
                "OBS_TERRENO",
                "OBSERVACION",
                "COMENTARIO",
                "NOTA",
                "REVISION_TERRENO",
                "DESCRIPCION_OBS",
            }:
                comentario_col = col
        if connector_col is not None:
            day_best: tuple[int, dict[int, int], int] | None = None
            for day_row in range(row, min(row + 6, ws.max_row) + 1):
                day_cols: dict[int, int] = {}
                for col in range(1, ws.max_column + 1):
                    day = _coerce_day(ws.cell(day_row, col).value)
                    if day is not None and day not in day_cols:
                        day_cols[day] = col
                if len(day_cols) >= 7:
                    score = len(day_cols)
                    if day_best is None or score > day_best[0]:
                        day_best = (score, day_cols, day_row)
            if day_best is not None:
                score, day_cols, day_row = day_best
                if best is None or score > best[0]:
                    best = (
                        score,
                        code_col,
                        matriz_col,
                        connector_col,
                        comentario_col,
                        day_cols,
                        day_row + 1,
                    )
    if best is None:
        raise ValueError(f"No se detectaron columnas mensuales validas en hoja '{ws.title}'.")
    _, code_col, matriz_col, connector_col, comentario_col, day_cols, data_start = best
    return code_col, matriz_col, connector_col, comentario_col, day_cols, data_start


def _lookup_monthly_rows(
    ws,
    connector_col: int,
    data_start: int,
) -> dict[str, int]:
    lookup: dict[str, int] = {}
    blank = 0
    max_scan_row = min(ws.max_row, data_start + 50000)
    min_col = connector_col
    max_col = connector_col
    for offset, values in enumerate(
        ws.iter_rows(
            min_row=data_start,
            max_row=max_scan_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    ):
        r = data_start + offset
        connector = safe_str(values[0] if values is not None else "")
        if not connector:
            blank += 1
            if blank >= 200:
                break
            continue
        blank = 0
        key = _normalize_connector(connector)
        if key not in lookup:
            lookup[key] = r
    return lookup


def _extraer_estados_desde_mensual(
    base_df: pd.DataFrame,
    ruta_programa_mensual: Path | str,
    fecha_desde: dt.date,
    fecha_hasta: dt.date,
) -> pd.DataFrame:
    output_cols = [
        "PUNTO",
        "TIPO",
        "DIA",
        "ESTADO",
        "COMENTARIO",
        "COMENTARIO_MENSUAL",
        "DESC_TAREA_2",
        "GRUPO_TRAB",
        "CODIGO",
        "MATRIZ",
        "ACTIVIDAD",
        "ORIGEN_HOJA",
        "ORIGEN_FILA",
    ]
    work_df = _filtrar_df_por_rango(base_df, fecha_desde, fecha_hasta)
    wb = openpyxl.load_workbook(Path(ruta_programa_mensual), data_only=True, read_only=False)
    try:
        ws = _select_monthly_sheet(wb, fecha_desde, fecha_hasta)
        if _is_sig_rpmo_sheet(ws.title):
            raise ValueError(
                "La hoja seleccionada del mensual es 'SIGRPMO', y no corresponde para clasificar cierre OT. "
                "Verifica que el archivo mensual sea el correcto y que tenga hoja de mes (ej: 'ABRIL 2026')."
            )
        _, _, connector_col, comentario_col, day_cols, data_start = _find_monthly_columns(ws)
        row_lookup = _lookup_monthly_rows(ws, connector_col, data_start)

        out_rows: list[dict[str, object]] = []
        total_input = 0
        total_grupo_filtrado = 0
        total_match_conector = 0
        total_match_dia = 0
        estados_detectados: dict[str, int] = {}
        for _, rec in work_df.iterrows():
            total_input += 1
            punto = safe_str(rec.get("PUNTO", ""))
            tipo = safe_str(rec.get("TIPO", ""))
            try:
                dia = int(rec.get("DIA"))
            except Exception:
                continue
            desc_tarea_2 = safe_str(rec.get("DESC_TAREA_2", ""))
            grupo_trab = safe_str(rec.get("GRUPO_TRAB", ""))
            if not _is_grupo_clasificable(grupo_trab):
                total_grupo_filtrado += 1
                continue
            matriz_terreno = tipo
            matriz_terreno_norm = _normalize_matriz_for_key(matriz_terreno)
            connector_candidates = []
            if desc_tarea_2 and matriz_terreno:
                connector_candidates.append(
                    _normalize_connector(f"{desc_tarea_2}/{matriz_terreno}")
                )
            if desc_tarea_2 and matriz_terreno_norm:
                connector_candidates.append(
                    _normalize_connector(f"{desc_tarea_2}/{matriz_terreno_norm}")
                )

            row = None
            for candidate in connector_candidates:
                row = row_lookup.get(candidate)
                if row is not None:
                    break
            if row is not None:
                total_match_conector += 1
            col = day_cols.get(dia)
            if row is not None and col is not None:
                total_match_dia += 1
            if row is None or col is None:
                continue
            estado = safe_str(ws.cell(row=row, column=col).value)
            estado_norm = normalize_text(estado)
            estados_detectados[estado_norm] = estados_detectados.get(estado_norm, 0) + 1
            if estado_norm not in ESTADOS_VALIDOS_MENSUAL:
                continue
            comentario_mensual = (
                safe_str(ws.cell(row=row, column=comentario_col).value)
                if comentario_col is not None
                else ""
            )
            out_rows.append(
                {
                    "PUNTO": punto,
                    "TIPO": tipo,
                    "DIA": dia,
                    "ESTADO": estado,
                    "COMENTARIO": comentario_mensual,
                    "COMENTARIO_MENSUAL": comentario_mensual,
                    "DESC_TAREA_2": desc_tarea_2,
                    "GRUPO_TRAB": grupo_trab,
                    "CODIGO": punto,
                    "MATRIZ": tipo,
                    "ACTIVIDAD": safe_str(rec.get("ACTIVIDAD", "")),
                    "ORIGEN_HOJA": safe_str(rec.get("ORIGEN_HOJA", "")),
                    "ORIGEN_FILA": rec.get("ORIGEN_FILA", ""),
                }
            )
    finally:
        wb.close()

    if not out_rows:
        estados_top = ", ".join(
            [f"{k or '<VACIO>'}: {v}" for k, v in sorted(estados_detectados.items(), key=lambda kv: kv[1], reverse=True)[:6]]
        )
        print(
            "[Etapa 3] Advertencia: no se encontraron estados R/NR/RR/S para clasificar "
            f"en hoja mensual '{ws.title}'. "
            f"Entrada={total_input}, grupo_no_permitido={total_grupo_filtrado}, "
            f"match_conector={total_match_conector}, match_conector_dia={total_match_dia}. "
            f"Estados detectados={estados_top or 'sin datos'}."
        )

    return pd.DataFrame(out_rows, columns=output_cols)


def _codigo_cierre_desde_estado(estado: object) -> str:
    est = normalize_text(estado)
    if est == "R":
        return "TT"
    if est in {"NR", "RR"}:
        return "MI"
    return ""


def _build_lookup_codigos_por_connector(
    df_clasificados: pd.DataFrame,
) -> dict[tuple[str, int], dict[str, str]]:
    lookup: dict[tuple[str, int], dict[str, str]] = {}
    for _, rec in df_clasificados.iterrows():
        desc_tarea_2 = safe_str(rec.get("DESC_TAREA_2", ""))
        matriz = safe_str(rec.get("TIPO", ""))
        dia_val = rec.get("DIA")
        try:
            dia = int(dia_val)
        except Exception:
            continue
        if not desc_tarea_2 or not matriz or not (1 <= dia <= 31):
            continue
        estado_norm = normalize_text(rec.get("ESTADO", ""))
        codigo = ""
        if estado_norm in {"R", "NR", "RR"}:
            codigo = _codigo_cierre_desde_estado(estado_norm)
        elif estado_norm == "S":
            codigo = safe_str(rec.get("CODIGO_CIERRE", ""))

        comentario_adiciona = safe_str(rec.get("COMENTARIO_MENSUAL", ""))
        if not comentario_adiciona:
            comentario_adiciona = safe_str(rec.get("COMENTARIO", ""))

        # Mantiene filas que al menos aportan comentario para "Comentario adiciona",
        # aunque no exista codigo de cierre (p.ej. estado S sin match en diccionario).
        if not codigo and not comentario_adiciona:
            continue

        matriz_norm = _normalize_matriz_for_key(matriz)
        candidates = [
            _normalize_connector(f"{desc_tarea_2}/{matriz}"),
            _normalize_connector(f"{desc_tarea_2}/{matriz_norm}"),
        ]
        for conector in candidates:
            lookup[(conector, dia)] = {
                "codigo": codigo,
                "comentario_adiciona": comentario_adiciona,
            }
    return lookup


def _actualizar_cierre_turno_planilla(
    path_cierre_turno_base: Path | str,
    df_clasificados: pd.DataFrame,
    output_path: Path | str,
) -> Path:
    lookup = _build_lookup_codigos_por_connector(df_clasificados)
    wb = openpyxl.load_workbook(Path(path_cierre_turno_base))
    try:
        for ws in wb.worksheets:
            if normalize_text(ws.title) == "SIGRPMO":
                continue
            day = _sheet_day_from_title(ws.title)
            if day is None:
                continue

            header_map = {
                normalize_column_name(ws.cell(row=1, column=c).value): c
                for c in range(1, ws.max_column + 1)
            }
            col_tarea = header_map.get(normalize_column_name("Descripción de tarea programada 2"))
            col_matriz = header_map.get(normalize_column_name("Matriz Terreno"))
            col_codigo = header_map.get(normalize_column_name("CODIGO DE CIERRE"))
            col_grupo_trab = header_map.get(normalize_column_name("Grupo trab"))
            col_comentario_adiciona = header_map.get(normalize_column_name("Comentario adiciona"))
            col_tarea = col_tarea or header_map.get(normalize_column_name("Descripcion de tarea programada 2"))
            if col_tarea is None or col_matriz is None:
                continue

            blank = 0
            for r in range(2, ws.max_row + 1):
                desc_tarea_2 = safe_str(ws.cell(row=r, column=col_tarea).value)
                matriz_terreno = safe_str(ws.cell(row=r, column=col_matriz).value)
                grupo_trab = (
                    safe_str(ws.cell(row=r, column=col_grupo_trab).value)
                    if col_grupo_trab is not None
                    else ""
                )
                if not desc_tarea_2 and not matriz_terreno:
                    blank += 1
                    if blank >= 200:
                        break
                    continue
                blank = 0

                if not desc_tarea_2 or not matriz_terreno:
                    continue
                if col_grupo_trab is not None and not _is_grupo_clasificable(grupo_trab):
                    continue

                matriz_norm = _normalize_matriz_for_key(matriz_terreno)
                conn_1 = _normalize_connector(f"{desc_tarea_2}/{matriz_terreno}")
                conn_2 = _normalize_connector(f"{desc_tarea_2}/{matriz_norm}")
                payload = lookup.get((conn_1, day)) or lookup.get((conn_2, day))
                if payload is None:
                    continue

                codigo = safe_str(payload.get("codigo", ""))
                comentario_adiciona = safe_str(payload.get("comentario_adiciona", ""))
                if col_codigo is not None and codigo:
                    ws.cell(row=r, column=col_codigo).value = codigo
                if col_comentario_adiciona is not None and comentario_adiciona:
                    ws.cell(row=r, column=col_comentario_adiciona).value = comentario_adiciona

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        return output_file
    finally:
        wb.close()


def ejecutar_etapa_1_limpieza_base(
    ruta_msewjo: str | Path,
    ruta_matriz_clasificacion: str | Path,
    carpeta_salida: str | Path,
    fecha_ancla_turno: dt.date | None = None,
) -> Dict[str, Path]:
    print("[Etapa 1] Cierre OT base desde MSEWJO...")
    stage_dir = _stage_dir(carpeta_salida, ETAPA_1_DIRNAME)

    if fecha_ancla_turno is None:
        fecha_ancla_turno = dt.date.today()
    turn_start, turn_end = get_turn_window(fecha_ancla_turno)
    cierre_turno_filename = (
        f"Cierre de OT Turno {turn_start.strftime('%d.%m')} al {turn_end.strftime('%d.%m')}.xlsx"
    )
    cierre_turno_path = stage_dir / cierre_turno_filename
    generar_cierre_turno_desde_msewjo(
        path_msewjo=ruta_msewjo,
        output_path=cierre_turno_path,
        path_diccionario=ruta_matriz_clasificacion,
        turn_start=turn_start,
        turn_end=turn_end,
        include_sin_fecha=True,
    )

    # Archivos tecnicos para etapas posteriores (clasificacion/merge).
    df_limpio = limpiar_msewjo(ruta_msewjo)
    msewjo_limpio_path = save_dataframe_to_excel(
        df_limpio, stage_dir / "msewjo_limpio_tecnico.xlsx"
    )

    df_cierre_base = generar_cierre_ot_base(df_limpio)
    cierre_base_path = save_dataframe_to_excel(
        df_cierre_base, stage_dir / "cierre_ot_base_tecnico.xlsx"
    )

    return {
        "cierre_ot_turno_base": cierre_turno_path,
        "msewjo_limpio_tecnico": msewjo_limpio_path,
        "cierre_ot_base_tecnico": cierre_base_path,
    }


def ejecutar_etapa_2_actualizar_mensual(
    ruta_programa_turno: str | Path,
    ruta_programa_mensual: str | Path,
    ruta_matriz_clasificacion: str | Path,
    carpeta_salida: str | Path,
    fecha_desde: str | dt.date | None = None,
    fecha_hasta: str | dt.date | None = None,
) -> Dict[str, Path]:
    print("[Etapa 2] Actualizacion programa mensual...")
    stage_dir = _stage_dir(carpeta_salida, ETAPA_2_DIRNAME)

    fecha_desde_parsed = _parse_fecha_usuario(fecha_desde, "fecha_desde")
    fecha_hasta_parsed = _parse_fecha_usuario(fecha_hasta, "fecha_hasta")
    if fecha_desde_parsed is not None and fecha_hasta_parsed is None:
        fecha_hasta_parsed = fecha_desde_parsed
    if fecha_hasta_parsed is not None and fecha_desde_parsed is None:
        fecha_desde_parsed = fecha_hasta_parsed
    if (
        fecha_desde_parsed is not None
        and fecha_hasta_parsed is not None
        and fecha_desde_parsed > fecha_hasta_parsed
    ):
        raise ValueError("fecha_desde no puede ser mayor que fecha_hasta.")

    mensual_actualizado_path = stage_dir / "programa_mensual_actualizado.xlsx"
    df_turno_aplicado = actualizar_programa_mensual(
        ruta_programa_turno,
        ruta_programa_mensual,
        mensual_actualizado_path,
        ruta_matriz_clasificacion,
        fecha_desde=fecha_desde_parsed,
        fecha_hasta=fecha_hasta_parsed,
    )
    interno_dir = stage_dir / ETAPA_2_INTERNO_DIRNAME
    interno_dir.mkdir(parents=True, exist_ok=True)
    registros_turno_aplicado_path = save_dataframe_to_excel(
        df_turno_aplicado, interno_dir / "registros_turno_aplicado.xlsx"
    )
    # Limpia artefactos antiguos de versiones previas para reducir ruido visual.
    legacy_registros = stage_dir / "registros_turno_aplicado.xlsx"
    if legacy_registros.exists():
        legacy_registros.unlink()
    legacy_duplicados = stage_dir / "duplicados.xlsx"
    if legacy_duplicados.exists():
        legacy_duplicados.unlink()
    legacy_no_cruzados = stage_dir / "no_cruzados.xlsx"
    if legacy_no_cruzados.exists():
        legacy_no_cruzados.unlink()
    legacy_actividades = stage_dir / "actividades_no_clasificadas.xlsx"
    if legacy_actividades.exists():
        legacy_actividades.unlink()
    legacy_diag = stage_dir / "diagnostico_no_cruzados_rapido.xlsx"
    if legacy_diag.exists():
        legacy_diag.unlink()
    no_cruzados_path = interno_dir / "no_cruzados.xlsx"
    actividades_no_clasificadas_path = interno_dir / "actividades_no_clasificadas.xlsx"
    diagnostico_no_cruzados_path = interno_dir / "diagnostico_no_cruzados_rapido.xlsx"

    return {
        "programa_mensual_actualizado": mensual_actualizado_path,
        "no_cruzados": no_cruzados_path,
        "actividades_no_clasificadas": actividades_no_clasificadas_path,
        "diagnostico_no_cruzados_rapido": diagnostico_no_cruzados_path,
    }


def ejecutar_etapa_3_clasificacion(
    ruta_matriz_clasificacion: str | Path,
    carpeta_salida: str | Path,
    ruta_cierre_base: str | Path | None = None,
    ruta_registros_turno_aplicado: str | Path | None = None,
    ruta_programa_mensual: str | Path | None = None,
    fecha_desde: str | dt.date | None = None,
    fecha_hasta: str | dt.date | None = None,
    ruta_programa_mensual_2: str | Path | None = None,
    fecha2_desde: str | dt.date | None = None,
    fecha2_hasta: str | dt.date | None = None,
) -> Dict[str, Path]:
    print("[Etapa 3] Clasificacion final cierre OT...")
    stage_dir = _stage_dir(carpeta_salida, ETAPA_3_DIRNAME)

    if ruta_programa_mensual is None:
        raise ValueError("Para etapa 3 debes indicar archivo de Programa Mensual.")

    fecha_desde_parsed = _parse_fecha_usuario(fecha_desde, "fecha_desde")
    fecha_hasta_parsed = _parse_fecha_usuario(fecha_hasta, "fecha_hasta")
    if fecha_desde_parsed is not None and fecha_hasta_parsed is None:
        fecha_hasta_parsed = fecha_desde_parsed
    if fecha_hasta_parsed is not None and fecha_desde_parsed is None:
        fecha_desde_parsed = fecha_hasta_parsed
    if (
        fecha_desde_parsed is not None
        and fecha_hasta_parsed is not None
        and fecha_desde_parsed > fecha_hasta_parsed
    ):
        raise ValueError("fecha_desde no puede ser mayor que fecha_hasta.")

    # El cierre base para etapa 3 es el archivo multihoja "Cierre de OT Turno ...".
    cierre_base_input: Path | None = None
    if ruta_cierre_base is not None:
        candidate = Path(ruta_cierre_base)
        if candidate.exists() and "Cierre de OT Turno" in candidate.name:
            cierre_base_input = candidate

    if cierre_base_input is None:
        cierre_base_input = _resolver_cierre_turno_base_path(
            carpeta_salida,
            fecha_desde=fecha_desde_parsed,
            fecha_hasta=fecha_hasta_parsed,
        )

    _eliminar_archivos_etapa_3_no_deseados(stage_dir)

    if not cierre_base_input.name.startswith("Cierre de OT Turno "):
        cierre_actualizado_path = stage_dir / "cierre_ot_turno_actualizado.xlsx"
    else:
        cierre_actualizado_path = stage_dir / cierre_base_input.name

    resumen_1 = etapa3_clasificacion(
        path_cierre_ot_base=cierre_base_input,
        path_programa_mensual=Path(ruta_programa_mensual),
        path_diccionario=Path(ruta_matriz_clasificacion),
        output_path=cierre_actualizado_path,
        hoja_mensual=None,
    )
    print(
        "[Etapa 3] Mensual 1 aplicado: "
        f"codigos={resumen_1.get('codigos_actualizados', 0)} "
        f"comentarios={resumen_1.get('comentarios_actualizados', 0)}"
    )

    mensual_usado_2_path: Path | None = None
    if ruta_programa_mensual_2 is not None and safe_str(ruta_programa_mensual_2):
        fecha2_desde_parsed = _parse_fecha_usuario(fecha2_desde, "fecha2_desde")
        fecha2_hasta_parsed = _parse_fecha_usuario(fecha2_hasta, "fecha2_hasta")
        if fecha2_desde_parsed is not None and fecha2_hasta_parsed is None:
            fecha2_hasta_parsed = fecha2_desde_parsed
        if fecha2_hasta_parsed is not None and fecha2_desde_parsed is None:
            fecha2_desde_parsed = fecha2_hasta_parsed
        if (
            fecha2_desde_parsed is not None
            and fecha2_hasta_parsed is not None
            and fecha2_desde_parsed > fecha2_hasta_parsed
        ):
            raise ValueError("fecha2_desde no puede ser mayor que fecha2_hasta.")

        mensual_usado_2_path = Path(ruta_programa_mensual_2)
        resumen_2 = etapa3_clasificacion(
            path_cierre_ot_base=cierre_actualizado_path,
            path_programa_mensual=mensual_usado_2_path,
            path_diccionario=Path(ruta_matriz_clasificacion),
            output_path=cierre_actualizado_path,
            hoja_mensual=None,
        )
        print(
            "[Etapa 3] Mensual 2 aplicado: "
            f"codigos={resumen_2.get('codigos_actualizados', 0)} "
            f"comentarios={resumen_2.get('comentarios_actualizados', 0)}"
        )

    result = {
        "cierre_ot_turno_actualizado": cierre_actualizado_path,
        "mensual_usado_1": Path(ruta_programa_mensual),
    }
    if mensual_usado_2_path is not None:
        result["mensual_usado_2"] = mensual_usado_2_path

    _eliminar_archivos_etapa_3_no_deseados(stage_dir)
    return result


def ejecutar_limpieza_y_base(
    ruta_msewjo: str | Path,
    ruta_matriz_clasificacion: str | Path,
    carpeta_salida: str | Path,
) -> Dict[str, Path]:
    return ejecutar_etapa_1_limpieza_base(
        ruta_msewjo=ruta_msewjo,
        ruta_matriz_clasificacion=ruta_matriz_clasificacion,
        carpeta_salida=carpeta_salida,
    )


def ejecutar_flujo(
    ruta_msewjo: str | Path,
    ruta_programa_turno: str | Path,
    ruta_programa_mensual: str | Path,
    ruta_matriz_clasificacion: str | Path,
    carpeta_salida: str | Path,
    fecha_desde: str | dt.date | None = None,
    fecha_hasta: str | dt.date | None = None,
) -> Dict[str, Path]:
    print("Iniciando procesamiento completo...")
    etapa_1 = ejecutar_etapa_1_limpieza_base(
        ruta_msewjo=ruta_msewjo,
        ruta_matriz_clasificacion=ruta_matriz_clasificacion,
        carpeta_salida=carpeta_salida,
    )
    etapa_2 = ejecutar_etapa_2_actualizar_mensual(
        ruta_programa_turno,
        ruta_programa_mensual,
        ruta_matriz_clasificacion,
        carpeta_salida,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    ruta_registros_turno_aplicado = (
        Path(carpeta_salida)
        / TEMP_DIRNAME
        / ETAPA_2_DIRNAME
        / ETAPA_2_INTERNO_DIRNAME
        / "registros_turno_aplicado.xlsx"
    )
    etapa_3 = ejecutar_etapa_3_clasificacion(
        ruta_matriz_clasificacion=ruta_matriz_clasificacion,
        carpeta_salida=carpeta_salida,
        ruta_cierre_base=etapa_1["cierre_ot_base_tecnico"],
        ruta_registros_turno_aplicado=ruta_registros_turno_aplicado,
    )

    result = {**etapa_1, **etapa_2, **etapa_3}
    print("Proceso completo finalizado.")
    for key, value in result.items():
        print(f"- {key}: {value}")
    return result


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Automatizador de cierre OT para monitoreo ambiental."
    )
    parser.add_argument("--msewjo", required=True, help="Ruta archivo MSEWJO")
    parser.add_argument("--turno", required=True, help="Ruta programa de monitoreo por turno")
    parser.add_argument("--mensual", required=True, help="Ruta programa de monitoreo mensual")
    parser.add_argument(
        "--matriz",
        required=True,
        help="Ruta Excel de diccionarios (hojas MATRIZ_TERRENO y COD_CIERRE)",
    )
    parser.add_argument("--output", required=True, help="Carpeta de salida")
    parser.add_argument(
        "--fecha-desde",
        required=False,
        help="Filtro opcional para etapa 2. Formato dd/mm/aaaa.",
    )
    parser.add_argument(
        "--fecha-hasta",
        required=False,
        help="Filtro opcional para etapa 2. Formato dd/mm/aaaa.",
    )
    return parser


if __name__ == "__main__":
    args = _build_parser().parse_args()
    ejecutar_flujo(
        ruta_msewjo=args.msewjo,
        ruta_programa_turno=args.turno,
        ruta_programa_mensual=args.mensual,
        ruta_matriz_clasificacion=args.matriz,
        carpeta_salida=args.output,
        fecha_desde=args.fecha_desde,
        fecha_hasta=args.fecha_hasta,
    )
